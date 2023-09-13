# -*- coding: utf-8 -*-
import math
import copy
import openpyxl
import pandas as pd
import pynput
import screeninfo
import time
import tkinter
from tkinter.font import Font
from mss import mss
from pynput.mouse import Controller, Button
import numpy as np
import pyautogui
import cv2
from paddleocr import PaddleOCR
import os
import re
from sklearn.cluster import KMeans
from collections.abc import Iterable
from tkinter.messagebox import *
from tkinter import ttk, RIGHT, Y, LEFT
from sklearn.metrics import pairwise_distances as cdist
from datetime import datetime
import Levenshtein

class eventKeyboard():
    """
    pressed(key) is a internal function for key board listener, please do not
        invoke!
    NOTICE!!!  pressed(key) would check whether main program is down every 10 sec,
        to make it continous working, call activeFlagSet(newFlag=1) to reactive the working status.


    StartListener(): start the listener, also safe for any unterminate listener.
    terminate(): end the listener

    statusGet(): return listener status, 0 for initiated, 1 for started, >=2 for specific keys were pressed,
        -1 for terminated.
    """

    def __init__(self) -> None:

        self.keyValue = -1
        self.timeIntervalStart = 0
        self.timeIntervalEnd = 0
        
        self.keysIntervalStart = 0
        self.keysIntervalEnd = 0
        
        self.activeFlag = -1
        self.keyPressed = pynput.keyboard.Listener(on_press = self.pressed, on_release = self.released)
        # self.counter = 0
        # self.shortcutThread = False
        
        self.__status = 0
        self.subFunctionDead = False
        self.shortcutFlag = False
        self.recordKey = []
        self.recordedshortcut = {"Key.alt_lz": 2} # Key.alt_lz is default shortcut

        self.keyboard_key_dict = {
    "\x01" : ['ctrl','a'],
    "\x02" : ['ctrl','b'],
    "\x03" : ['ctrl','c'],
    "\x04" : ['ctrl','d'],
    "\x05" : ['ctrl','e'],
    "\x06" : ['ctrl','f'],
    "\x07" : ['ctrl','g'],
    "\x08" : ['ctrl','h'],
    "\t"   : ['ctrl','i'],
    "\n"   : ['ctrl','j'],
    "\x0b" : ['ctrl','k'],
    "\x0c" : ['ctrl','l'],
    "\r"   : ['ctrl','m'],
    "\x0e" : ['ctrl','n'],
    "\x0f" : ['ctrl','o'],
    "\x10" : ['ctrl','p'],
    "\x11" : ['ctrl','q'],
    "\x12" : ['ctrl','r'],
    "\x13" : ['ctrl','s'],
    "\x14" : ['ctrl','t'],
    "\x15" : ['ctrl','u'],
    "\x16" : ['ctrl','v'],
    "\x17" : ['ctrl','w'],
    "\x18" : ['ctrl','x'],
    "\x19" : ['ctrl','y'],
    "\x1a" : ['ctrl','z'],
    "\x1f" : ['ctrl','shift','-'],
    '<186>'  : ['ctrl',';'],
    "<187>"  : ['ctrl','='],
    "<189>"  : ['ctrl','-'],
    "<192>"  : ['ctrl','`'],
    "<222>"  : ['ctrl',r"'"],
    "<48>"   : ['ctrl','0'],
    "<49>"   : ['ctrl','1'],
    "<50>"   : ['ctrl','2'],
    "<51>"   : ['ctrl','3'],
    "<52>"   : ['ctrl','4'],
    "<53>"   : ['ctrl','5'],
    "<54>"   : ['ctrl','6'],
    "<55>"   : ['ctrl','7'],
    "<56>"   : ['ctrl','8'],
    "<57>"   : ['ctrl','9'],
    "~"    : ['shift', '`'],
    "!"    : ['shift', '1'],
    "@"    : ['shift', '2'],
    "#"    : ['shift', '3'],
    "$"    : ['shift', '4'],
    "%"    : ['shift', '5'],
    "^"    : ['shift', '6'],
    "*"    : ['shift', '7'],
    "("    : ['shift', '8'],
    ")"    : ['shift', '9'],
    "_"    : ['shift', '-'],
    "+"    : ['shift', '='],
    ":"    : ['shift', ';'],
    "\'"   : ['shift', "'"],
    "<"    : ['shift', ","],
    "{"    : ['shift', "["],
    "}"    : ['shift', "]"],
    "|"    : ['shift', "\\"],
    "?"    : ['shift', "/"],
}

    def pressed(self,key):
        if self.subFunctionDead:
            return True
        self.keyPressReset = time.time()
        try:
            self.recordKey.append("{}".format(key.char))
        except:
            self.recordKey.append("{}".format(key))
        # print(self.recordKey)
        
        if len(self.recordKey) == 2:
            # print("recordKey",self.recordKey[0])
            if self.recordKey[0] == "Key.ctrl_l" or self.recordKey[0] == "Key.ctrl_r":
                # print("---------------------")
                if self.recordKey[1] in self.keyboard_key_dict.keys():
                    # print(key)
                    # print("The recordKey is " + str(self.keyboard_key_dict[self.recordKey[1]]))
                    # print("Record the shortcut key.")
                    self.recordKey[1] = copy.copy(self.keyboard_key_dict[self.recordKey[1]][1])

            elif self.recordKey[0] == "Key.shift" or self.recordKey[0] == "Key.shift_r":
                # print("+++++++++++++++++++=")
                if self.recordKey[1] in self.keyboard_key_dict.keys():
                    # print("The recordKey is " + str(self.keyboard_key_dict[self.recordKey[1]]))
                    # print("Record the shortcut key.")
                    self.recordKey[1] = copy.copy(self.keyboard_key_dict[self.recordKey[1]][1])

            if self.shortcutFlag == False:
                # print("call target functions, not finished yet.")
                    # else:
                    #     # print(type(self.recordKey[0]))
                    #     print("The recordKey is " + self.recordKey[0] + " + " + self.recordKey[1])
                    #     print("Record the shortcut key.")
                    #     print("sssssssssssssssssssssss")
                    tem = self.recordKey[0] + self.recordKey[1]
                    if tem in self.recordedshortcut.keys():
                        print(self.recordedshortcut[tem])
                        self.__status = self.recordedshortcut[tem] # Change status if user triggered a shortcut
        
        if time.time() - self.timeIntervalStart > 10.0:
            if self.activeFlag == -1: # check if main thread is active, if not, kill sub thread
                self.subFunctionDead = True
                self.__status = -1
                self.recordKey.clear()
                
                return False
            else:
                self.activeFlag = -1
                self.timeIntervalStart = time.time()

    def released(self,key):
        if self.subFunctionDead:
            return True
        self.keyPressReset = time.time()
        try:
            tem = str(key.char)
        except:
            tem = str(key)
            
        if tem in self.keyboard_key_dict.keys():
            tem=self.keyboard_key_dict[tem][1]
            
        if self.shortcutFlag and len(self.recordKey) == 2: # record shortcut
            if len(list(self.recordedshortcut.keys())) == 0:
                self.recordedshortcut[self.recordKey[0] + self.recordKey[1]] = 2
            else:
                if (self.recordKey[0] + self.recordKey[1]) not in self.recordedshortcut.keys():
                    self.recordedshortcut[self.recordKey[0] + self.recordKey[1]] = \
                        self.recordedshortcut[list(self.recordedshortcut.keys())[-1]] + 1
            # print(self.recordedshortcut)
            self.shortcutFlag = False
        
        # print("Released!!!",self.recordKey)
        # try:
        #     self.recordKey.remove("{}".format(key.char))
        # except:
        if len(self.recordKey) > 0:
            try:
                self.recordKey.remove("{}".format(tem))
            except:
                self.recordKey.clear()
        
        if time.time() - self.timeIntervalStart > 10.0:
            if self.activeFlag == -1: # check if main thread is active, if not, kill sub thread
                self.subFunctionDead = True
                self.__status = -1
                self.recordKey.clear()
                return False
            else:
                self.activeFlag = -1
                self.timeIntervalStart = time.time()
        # if key == pynput.keyboard.Key.esc:
        #     return False

        
    def StartListener(self) -> None:

        if self.__status == -1:
            self.terminate()
            self.keyPressed = pynput.keyboard.Listener(on_press=self.pressed, on_release=self.released)
            
        self.__status = 1
        self.keyPressed.start()
        self.timeIntervalStart = time.time()
        self.begin = time.time() - 5
        self.keyPressReset = time.time()
        self.recordKey = []
        self.subFunctionDead = False

    def terminate(self) -> None:
        self.keyPressed.stop()

    def keyGet(self) -> str:
        return self.keyValue

    def activeFlagSet(self, flag) -> None:
        end = time.time()
        dif = end - self.begin
        if self.__status == -1:
            self.StartListener()
        elif dif >= 10.0:
            self.activeFlag = flag
            self.begin = time.time()

    def statusGet(self) -> int:
        if time.time() - self.keyPressReset > 1.0:
            self.__status = 1
            
            return 1
        result = self.__status
        if self.__status > 1:
            self.__status = 1
            
        return result


class eventMouse():
    """
    clicked(x,y) and moving(x,y) are internal functions for key board listener, please do not
        invoke!
    NOTICE!!!  clicked(x,y) and moving(x,y) would check whether main program is down every 10 sec,
        to make it continous working, call activeFlagSet(newFlag=1) to reactive the working status.

    StartListener(): start the listeners.
    terminate(): end the listeners.

    mouseGet(side): if side can be "left" or "right",return the x,y coordinate for last time mouse clicked

    motionGet(): return the x,y coordinate for last time mouse moved
    """

    def __init__(self) -> None:
        self.timeIntervalStart = 0
        self.timeIntervalEnd = 0

        self.activeFlag1 = -1
        self.activeFlag2 = -1
        
        self.__status1 = 0
        self.__status2 = 0

        self.DetectedMouseXPos = -1
        self.DetectedMouseYPos = -1

        self.DetectedRightMouseXPos = -1
        self.DetectedRightMouseYPos = -1

        self.timeIntervalStartMotion = 0
        self.timeIntervalEndMotion = 0
        self.MotionMouseXPos = -1
        self.MotionMouseYPos = -1

        # keyPressed.stop()
        self.mouseClicked = pynput.mouse.Listener(on_click=self.clicked)
        self.mouseMove = pynput.mouse.Listener(on_move=self.moving)

        # mouseClicked.join()

    def moving(self, x, y):
        self.MotionMouseXPos = x
        self.MotionMouseYPos = y

        self.timeIntervalEndMotion = time.time()
        if self.timeIntervalEndMotion - self.timeIntervalStartMotion > 10.0:
            # print("Times Up")
            if self.activeFlag2 == -1:
                # print("exit")
                self.__status2 = -1
                return False
            self.timeIntervalStartMotion = time.time()
            self.activeFlag2 = -1

    def clicked(self, x, y, button, pressed):
        self.mouseClickReset = time.time()
        
        if pressed and button.name == "left":
            self.DetectedMouseXPos = x
            self.DetectedMouseYPos = y
            # print(self.DetectedMouseXPos,self.DetectedMouseYPos)

        if pressed and button.name == "right":
            self.DetectedRightMouseXPos = x
            self.DetectedRightMouseYPos = y
            # print(self.DetectedRightMouseXPos,self.DetectedRightMouseYPos)

        self.timeIntervalEnd = time.time()

        # print(self.timeIntervalEnd-self.timeIntervalStart)
        if self.timeIntervalEnd - self.timeIntervalStart > 10.0:
            #             print("Times Up")
            if self.activeFlag1 == -1:
                #                 print("exit")
                self.__status1 = -1
                return False
            self.timeIntervalStart = time.time()
            self.activeFlag1 = -1
            # if x==0 or y ==0:
            #     self.detectFlag=False
            # self.terminate()
            # return True

        # return True

    def StartListener(self) -> None:
        
        if self.__status1 == -1 or self.__status2 == -1:
            self.terminate()
            self.mouseClicked = pynput.mouse.Listener(on_click=self.clicked)
            self.mouseMove = pynput.mouse.Listener(on_move=self.moving)
        
        self.timeIntervalStart = time.time()
        self.mouseClicked.start()

        self.timeIntervalStartMotion = time.time()
        self.mouseMove.start()

        self.begin = time.time() - 5
        self.mouseClickReset = time.time()

    def terminate(self) -> None:
        self.mouseClicked.stop()
        self.mouseMove.stop()

    def mouseGet(self, side) -> int:
        if time.time() - self.mouseClickReset > 1.0:
            self.DetectedMouseXPos = -1
            self.DetectedMouseYPos = -1

            self.DetectedRightMouseXPos = -1
            self.DetectedRightMouseYPos = -1
            
            return -1, -1
        
        if side == "left":
            return self.DetectedMouseXPos, self.DetectedMouseYPos
        elif side == "right":
            return self.DetectedRightMouseXPos, self.DetectedRightMouseYPos

    def motionGet(self) -> int:
        return self.MotionMouseXPos, self.MotionMouseYPos

    def activeFlagSet(self, newFlag) -> None:
        end = time.time()
        dif = end - self.begin
        
        if self.__status1 == -1 or self.__status2 == -1:
            self.StartListener()
        elif dif >= 10:
            self.activeFlag1 = newFlag  # 检测点击的flag
            self.activeFlag2 = newFlag  # 检测移动的flag
            self.begin = time.time()
            
    def restart(self):
        self.DetectedMouseXPos = -1
        self.DetectedMouseYPos = -1

        self.DetectedRightMouseXPos = -1
        self.DetectedRightMouseYPos = -1

        self.MotionMouseXPos = -1
        self.MotionMouseYPos = -1


class windowsUI():
    """
    class parameters:
        `override`: self defined tk windows, only set true when using screenShot mode or message box mode
        `alpha`: visibility of root window
        `bgColor`: background color, only root windows
        `screenShot`: mode flag, 1 for screenShot mode, 2 for main window mode, 3 for child windows on screenShot mode
        width,height: size of root window
        positionX,positionY: x,y for top left of window
        `listener`: mouse listener

        !!! Flag summary:
        `self.statusID`:
            singel digit: the function ID for current view windows/panels
            ten digit: the view ID for current window/panel
            hundred digit: special digit repersent the main program status
            thousand digit: the ID of window/panel
            Example:
                1100:Start up screen shot
                    first "1" means root window
                    second "1" means the button has pressed and wait for event function to handle
                    first "0" means first view of root window
                    second "0" means the the button to Start up screen shot is on first view of root window

            detials:
            -1: standby flag
            "1100": record user behaviour
            "1103": Start up screen shot
            "2000": Screen shot mode
            "1010": behaviour recording panel
            "1020": click setting panel
            "1030-1099": choose available button
            "3000": show result
            "4000": execute command
            "5000": standby config
            "6000": loop config
            "7000": scorlling
    """

    def __init__(self, override=False, alpha=0.8, bgColor="black", screenShot=-1, \
                 width=-1, height=-1, positionX=0, positionY=0, listener: eventMouse = None) -> None:
        # self.__timeList=[time.time(),0]
        self.test = False # enable test code
        self.testNum = 0
        
        self.listener = eventMouse()
        self.keyBoardInterrupt = eventKeyboard()
        self.screenShot = screenShot
        self.screen = screeninfo.get_monitors()[0]
        # print("22222222222222222222222333333333333333333333")
        self.mainPanelButtons = ({"Recognition Area Record": [1100], "Execute Recorded": [1101],\
            "Setting": [1102], "next page": [1103], "testRecord": [1104]},
                {"Mouse Click": [1110], "Mouse Hold": [1111], "Mouse Move": [1112], \
                    "Scorlling": [1113],"Text Recognize": [1114], "Loop": [1115], \
                        "Standby": [1116], "form the data": [1117],"Save record": [1118], "Back to home page": [1119]},
                {"Back to home page": [4100]}, 
                {"record by coordinate": [1120],"click on a button": [1121],"back": [1122]}, 
                {"back": [1130]}, {"wait by time counts": [5100], "wait until specific text show up": [5101], \
                    "wait until spcific icon or button show up(recommend)": [5102], "back": [5103]}, \
                {"confirm": [5110], "Cancel": [5111]}, {"back": [5120]}, {"Cancel": [6100]}, \
                {"a Piece of text": [6110], "list of text": [6111], "Cancel": [6112]}, \
                {"Cancel": [6120], "Set up new model": [6121]}, {"Cancel": [6300], \
                    "Set up new model": [6301]}, {"No": [5120]}, {"Begin of the text": [6400], \
                        "End of the text": [6401]}, \
                {"click on text boxes(centre)": [6410],"click on a button": [6411],"Text retrieval": [6412],\
                    "DONE": [6413]}, {}, {"Area text recognize(Faster,lower accuracy)": [6500], \
                        "text matching(Slower,higher accuracy)": [6501]}, \
                {"Cancel":[6520], "member information":[6521], "clans battles":[6522]}, \
                {"Cancel":[6540], "Confirm":[6541]}) # 18
        self.mainPanelInput = ([],[],[],[],[],[],["Hours", "Minutes", "Seconds"],[],\
            ["Loop time(-1 for infinite loop and will only stop by user interrupt)"], \
                [],[],[],[],[],[],[],[], [], []) # 18
        self.mainPanelLabel = ([],[],[],[],[],[],[],[],[],[],["Please choose one of model to \
find out text boxes"], ["Please choose one of model to filter out error location"],\
    ["Is there any icon or UI could easily be consider as text?"], ["Where the icon locate at?"], [], [], [], \
        ["Which form would you like to save your data?"], \
        ["Please correspond your data with correct title, this form only show the first line"]) # 18
        
        self.mainPanelCombobox = ([],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],\
        [[6, "昵称","标签","加入","最近退出","差值","常驻认证T/F"]]) # 18
        
        self.description = {2010:"leftClick", 1131: "clickOnButton", 5110:"timeWait", 5121: "iconDetection", \
            6101:"loopController", 6543:"dataForm"}
        
        self.formDataDesicion = ["--==//binary operation--==//", "--==//time--==//", "--==//special condition--==//"]
        self.timeCondition = ["Current time YY:MM:DD:H:M:S", "Time appear on list YY:MM:DD:H:M:S", \
            "Time disappear from list YY:MM:DD:H:M:S"]
        self.binaryCalculation = ["+", "-", "x", "÷"]
        self.specialCondition = ["=", ">", "<", ">=", "<="]
        self.mapForm = dict()
        self.formConditionSelection = []
        self.specialData = 0
        
        self.displacementMouse = 0
        self.displacementCount = 0
        
        self.temFileName = ""
        
        
        # Store the buttons on main Panel and their status ID
        self.currentButton: list[tkinter.Button] = []
        self.currentLabel = []
        self.currentInput: list[tkinter.Entry] = []
        
        self.currentButtonSubWin: list[tkinter.Button] = []
        self.currentLabelSubWin = []
        
        self.currentOtherWidgets = []
        

        
        self.loopCounter = ([],[],[],[]) # index of loop back to, loop time, if finished loop

        self.statusID = 1000  # globel status flag
        self.__subWindows = None  # store windows created by event function Start()
        self.messageBox = None
        self.screenShotor = None

        self.__rec = []  # store rectangle in canvas
        self.recoredArea = [] # store recorded screen shot area
        self.textBoxes = [] # store recognized text boxes
        self.textAreUpperBound = self.screen.height
        self.textAreLowerBound = 0
        
        self.sequencialCommand = []
        
        self.x = -10  # store mouse click coordinations
        self.y = -10
        
        self.dbManagement = edit_excel()
        self.ocr = OCRController(self.dbManagement.currentPath, self.dbManagement.OCRModelDataLoader())

        self.testModelID = None
        self.testScrollingArea = {"top": 0, "left": 0, "width": 0, "height": 0}
        self.xRight = -1  # store mouse move coordinations
        self.yRight = -1

        # self.__loopTime = 0  # use to count the time, 0.1s every loop
        self.__counter = 0  # status id for drawer function
        self.counter = -1
        self.__root = tkinter.Tk()
        self.__root.iconbitmap("agent.ico")   # 更改窗口图标

        self.windowSize = {"root": [width, height, positionX, positionY], "screenShoter": \
            [0, 0, 0, 0], "record": [0,0,0,0]}  # all type of window size

        self.width = width  # current use width and height
        self.height = height
        self.bgColor = bgColor
        self.alpha = alpha

        
        self.scroller = None
        self.mylistBox = None
        
        self.imageFilterOut = None
        self.indexFilterOut = None
        
        self.currentColumnNum = -1
        self.temForm = dict()
        
        
        # testButton = tkinter.Button(self.__root, text="tester")
        # font = Font(testButton["font"])  # get font information
        # self.grid = font.metrics("linespace")  # calculate hieght and weidth by font information
        # # lineWidth = font.measure(x)
        # testButton.destroy()
        self.positionX = positionX
        self.positionY = positionY
        
        self.IOController = mouse_control()
        self.userInteraction = UserBehaviourController(self.IOController, self.listener, self.ocr,\
            self.dbManagement, self)
        
        self.recordPointer = len(self.userInteraction.recoredBehaviours)
        self.executePointer = 0

        if self.screenShot == 1:
            self.__num = 6
            self.canvasPlace()

            if width == -1:
                self.width = self.__root.winfo_screenwidth()
            if height == -1:
                self.height = self.__root.winfo_screenheight()

            # self.__root.overrideredirect(override)
            # self.__root.attributes("-alpha", alpha)
        elif self.screenShot == 2:

            # temx=200*(1280/self.screen.width)
            # temy=70*(720/self.screen.height)
            # print(temx,temy)
            self.__root.title("Main panel")
            if width == -1:
                self.width = int(self.__root.winfo_screenwidth() / 5)
            if height == -1:
                self.height = int(self.width * 16 / 10)
                
            self.windowSize["root"][0] = self.width
            self.windowSize["root"][1] = self.height
            # print(self.width,self.height)
            self.layOutController()
        self.__root.overrideredirect(override)
        self.__root.attributes("-alpha", alpha)

        self.__root.geometry("{0}x{1}+{2}+{3}" \
                             .format(self.width, self.height, positionX, positionY))
        self.__root.configure(bg=bgColor)

        self.__root.resizable(0, 0)

        if self.listener != None:
            self.listener.StartListener()
        if self.screenShot != 1:
            self.keyBoardInterrupt.StartListener()  # listener to detect keyboard shortcut
        self.keeper()

        self.__root.mainloop()

    def layOutController(self, mode="root", view=0, lastmode = "None", windows = "root", align = "downward") -> None:
        if windows == "root":
            counter = 0
            buttonNum = len(self.mainPanelButtons[view].keys())
            buttonNum += len(self.mainPanelInput[view])
            buttonNum += len(self.mainPanelLabel[view])
            buttonStatus = [tkinter.NORMAL, tkinter.DISABLED]
            buttonStatusSetup = []
            if mode == "root":
                self.widgetsCleaner(self.currentButton, self.currentLabel, self.currentButtonSubWin, \
                    self.currentLabelSubWin, self.scroller, self.mylistBox, self.currentInput, self.currentOtherWidgets)
                if self.__subWindows != None:
                    self.__subWindows.destroy()
                    self.__subWindows = None
                    
                self.width, self.height, self.positionX, self.positionY = self.windowSize[mode]


                self.__root.resizable(0, 0)

            elif mode == "record":
                self.widgetsCleaner(self.currentButton, self.currentLabel, self.currentInput, \
                    self.currentOtherWidgets)
                print(self.currentButton)
                # for x in range(len(self.currentButton)):
                #     self.currentButton[x].destroy()

                self.windowSize[lastmode] = []  # back up the size of root window
                self.windowSize[lastmode].append(self.width)
                self.windowSize[lastmode].append(self.height)
                self.windowSize[lastmode].append(self.positionX)
                self.windowSize[lastmode].append(self.positionY)

                self.width = int(self.__root.winfo_screenwidth() / 5)
                self.height = int(self.width * 2 / 1)
                self.positionX = 50
                self.positionY = 50


                self.__root.resizable(0, 0)

            elif mode == "executeList" or mode == "clickRecord" or mode == "UIClick" or mode == "standbyConfig"\
                or mode == "textRecognizeConfig" or mode == "extraSteps" or mode == "extraStepsUIClick"\
                    or mode == "TextRetrievalMode":
                
                self.widgetsCleaner(self.currentButton, self.currentLabel, self.currentInput, \
                    self.currentOtherWidgets)

                self.windowSize[lastmode] = []  # back up the size of last window
                self.windowSize[lastmode].append(self.width)
                self.windowSize[lastmode].append(self.height)
                self.windowSize[lastmode].append(self.positionX)
                self.windowSize[lastmode].append(self.positionY)

                self.width = int(self.__root.winfo_screenwidth() / 5)
                self.height = int(self.width * 2 / 1)
                self.positionX = 50
                self.positionY = 50
                
            elif mode == "timeWaitConfig":
                self.widgetsCleaner(self.currentButton, self.currentLabel, self.currentInput, \
                    self.currentOtherWidgets)

                self.windowSize[lastmode] = []  # back up the size of last window
                self.windowSize[lastmode].append(self.width)
                self.windowSize[lastmode].append(self.height)
                self.windowSize[lastmode].append(self.positionX)
                self.windowSize[lastmode].append(self.positionY)

                self.width = int(self.__root.winfo_screenwidth() / 5)
                self.height = int(self.width * 13 / 10)
                self.positionX = 50
                self.positionY = 50
                
                for x in self.mainPanelInput[view]:# place input box
                    counter += 1
                    entry_var = tkinter.StringVar()
                    entry_var.set(x)
                    self.currentInput.append(tkinter.Entry(self.__root,width=int(self.width/12),\
                        textvariable=entry_var))
                    font = Font(font=self.currentInput[-1]["font"])  # get font information
                    lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                    lineWidth = font.measure("  "*int(self.width/12))
                    print(self.width)
                    print(lineHeight, lineWidth)
                    # print(lineHeight,lineWidth)
                    # print()
                    #
                    self.currentInput[-1].place(x=(self.width - lineWidth) / 2,
                                                 y=counter * self.height / (buttonNum + 1) - lineHeight / 2)

            elif mode == "loopConfig":
                self.widgetsCleaner(self.currentButton, self.currentLabel, self.currentInput, \
                    self.currentOtherWidgets)

                self.windowSize[lastmode] = []  # back up the size of last window
                self.windowSize[lastmode].append(self.width)
                self.windowSize[lastmode].append(self.height)
                self.windowSize[lastmode].append(self.positionX)
                self.windowSize[lastmode].append(self.positionY)

                self.width = int(self.__root.winfo_screenwidth() / 5)
                self.height = int(self.width * 2 / 1)
                self.positionX = 50
                self.positionY = 50
                
                self.mainPanelButtons[view].clear()
                self.mainPanelButtons[view]["Cancel"] = [6100]
                
                buttonStatusSetup.append(0)
                for x in range(len(self.loopCounter)):
                    if len(self.loopCounter[x]) == 0:
                        self.mainPanelButtons[view]["Start Loop " + chr(65 + x)] = [6100 + x + 1]
                        buttonStatusSetup.append(0)
                    else:
                        if self.loopCounter[x][2]:
                            buttonStatusSetup.append(1)
                        else:
                            buttonStatusSetup.append(0)
                        self.mainPanelButtons[view]["Stop Loop " + chr(65 + x)] = [6100 + x + 1]
                buttonNum = len(self.mainPanelButtons[view].keys()) + len(self.mainPanelInput[view]) \
                    + len(self.mainPanelLabel[view])
                
                for x in self.mainPanelInput[view]:# place input box
                    counter += 1
                    entry_var = tkinter.StringVar()
                    entry_var.set(x)
                    self.currentInput.append(tkinter.Entry(self.__root,width=int(self.width/12),\
                        textvariable=entry_var))
                    font = Font(font=self.currentInput[-1]["font"])  # get font information
                    lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                    lineWidth = font.measure("  "*int(self.width/12))
                    print(self.width)
                    print(lineHeight, lineWidth)
                    # print(lineHeight,lineWidth)
                    # print()
                    #
                    self.currentInput[-1].place(x=(self.width - lineWidth) / 2,
                                                 y=counter * self.height / (buttonNum + 1) - lineHeight / 2)
                
            elif mode == "textBoxesModelling":
                self.widgetsCleaner(self.currentButton, self.currentLabel, self.currentInput, \
                    self.currentOtherWidgets)

                self.windowSize[lastmode] = []  # back up the size of last window
                self.windowSize[lastmode].append(self.width)
                self.windowSize[lastmode].append(self.height)
                self.windowSize[lastmode].append(self.positionX)
                self.windowSize[lastmode].append(self.positionY)

                self.width = int(self.__root.winfo_screenwidth() / 5)
                self.height = int(self.width * 2 / 1)
                self.positionX = 50
                self.positionY = 50
                
                for x in self.mainPanelLabel[view]:# place buttons
                    counter += 1
                    self.currentLabel.append(tkinter.Label(self.__root, text=x, fg="white", bg="black"))

                    font = Font(font=self.currentLabel[-1]["font"])  # get font information
                    lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                    lineWidth = font.measure(x)
                    # print(lineHeight,lineWidth)
                    # print()
#       
                    self.currentLabel[-1].place(x=(self.width - lineWidth) / 2,
                                                 y=counter * self.height / (buttonNum + 1) - lineHeight / 2)

            elif mode == "textSeeker":
                self.widgetsCleaner(self.currentButton, self.currentLabel, self.currentInput, \
                    self.currentOtherWidgets)

                self.windowSize[lastmode] = []  # back up the size of last window
                self.windowSize[lastmode].append(self.width)
                self.windowSize[lastmode].append(self.height)
                self.windowSize[lastmode].append(self.positionX)
                self.windowSize[lastmode].append(self.positionY)

                self.width = int(self.__root.winfo_screenwidth() / 5)
                self.height = int(self.width * 2 / 1)
                self.positionX = 50
                self.positionY = 50
                
                for x in self.mainPanelLabel[view]:# place buttons
                    counter += 1
                    self.currentLabel.append(tkinter.Label(self.__root, text=x, fg="white", bg="black"))

                    font = Font(font=self.currentLabel[-1]["font"])  # get font information
                    lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                    lineWidth = font.measure(x)
                    # print(lineHeight,lineWidth)
                    # print()
#       
                    self.currentLabel[-1].place(x=(self.width - lineWidth) / 2,
                                                 y=counter * self.height / (buttonNum + 1) - lineHeight / 2)

            elif mode == "iconFilterOut":
                self.widgetsCleaner(self.currentButton, self.currentLabel, self.currentInput, \
                    self.currentOtherWidgets)

                self.windowSize[lastmode] = []  # back up the size of last window
                self.windowSize[lastmode].append(self.width)
                self.windowSize[lastmode].append(self.height)
                self.windowSize[lastmode].append(self.positionX)
                self.windowSize[lastmode].append(self.positionY)

                self.width = int(self.__root.winfo_screenwidth() / 5)
                self.height = int(self.width * 2 / 1)
                self.positionX = 50
                self.positionY = 50
                
                for x in self.mainPanelLabel[view]:# place buttons
                    counter += 1
                    self.currentLabel.append(tkinter.Label(self.__root, text=x, fg="white", bg="black"))

                    font = Font(font=self.currentLabel[-1]["font"])  # get font information
                    lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                    lineWidth = font.measure(x)
                    # print(lineHeight,lineWidth)
                    # print()
#       
                    self.currentLabel[-1].place(x=(self.width - lineWidth) / 2,
                                                 y=counter * self.height / (buttonNum + 1) - lineHeight / 2)

            elif mode == "positionSelect":
                self.widgetsCleaner(self.currentButton, self.currentLabel, self.currentInput, \
                    self.currentOtherWidgets)

                self.windowSize[lastmode] = []  # back up the size of last window
                self.windowSize[lastmode].append(self.width)
                self.windowSize[lastmode].append(self.height)
                self.windowSize[lastmode].append(self.positionX)
                self.windowSize[lastmode].append(self.positionY)

                self.width = int(self.__root.winfo_screenwidth() / 5)
                self.height = int(self.width * 2 / 1)
                self.positionX = 50
                self.positionY = 50
                
                for x in self.mainPanelLabel[view]:# place buttons
                    counter += 1
                    self.currentLabel.append(tkinter.Label(self.__root, text=x, fg="white", bg="black"))

                    font = Font(font=self.currentLabel[-1]["font"])  # get font information
                    lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                    lineWidth = font.measure(x)
                    # print(lineHeight,lineWidth)
                    # print()
#       
                    self.currentLabel[-1].place(x=(self.width - lineWidth) / 2,
                                                 y=counter * self.height / (buttonNum + 1) - lineHeight / 2)

            elif mode == "memberInfoForm":
                self.widgetsCleaner(self.currentButton, self.currentLabel, self.currentInput, \
                    self.currentOtherWidgets)

                self.windowSize[lastmode] = []  # back up the size of last window
                self.windowSize[lastmode].append(self.width)
                self.windowSize[lastmode].append(self.height)
                self.windowSize[lastmode].append(self.positionX)
                self.windowSize[lastmode].append(self.positionY)

                self.width = int(self.__root.winfo_screenwidth() / 1.25)
                self.height = int(self.__root.winfo_screenheight() / 4)
                self.positionX = 50
                self.positionY = 50
                
                rowNumber = len(self.mainPanelCombobox[view])
                columnNum = self.mainPanelCombobox[view][0][0]
                com = ttk.Combobox(self.__root)     # #创建下拉菜单
                com.place(x = 0, y = 0)     # #将下拉菜单绑定到窗体
                self.__root.update()
                com["value"] = ("test1", "test2", "test3")    # #给下拉菜单设定值
                com.current(2)    # #设定下拉菜单的默认值为第3个
                width = com.winfo_width()
                height = com.winfo_height()
                com.destroy()
                counter = 0
                # print(rowNumber)
                for x in range(rowNumber):# place combobox as form
                    
                    for y in range(self.mainPanelCombobox[view][x][0]):
                        counter = y
                        self.currentOtherWidgets.append(ttk.Combobox(self.__root))
                        if x > 0:
                            self.currentOtherWidgets[-1]["value"] = self.mainPanelCombobox[view][x][1:] + self.formDataDesicion
                        else:
                            self.currentOtherWidgets[-1]["value"] = self.mainPanelCombobox[view][x][1:]
                        self.currentOtherWidgets[-1].current(y)
                        # print(self.height)
                        # print(self.height - height*rowNumber\
                        #         - 1*(rowNumber-1) / 2 + x*(height+1))
                        # print(height)
                        self.currentOtherWidgets[-1].place(x=(self.width - width*columnNum - \
                            1*(columnNum-1)) / 2 + y*(width + 1), y=(self.height - height*rowNumber\
                                - 1*(rowNumber-1)) / 2 + x*(height+1))

                    while len(self.currentOtherWidgets) % columnNum != 0:
                        counter += 1
                        self.currentOtherWidgets.append(ttk.Combobox(self.__root))
                        self.currentOtherWidgets[-1]["value"] = self.formDataDesicion
                        self.currentOtherWidgets[-1].place(x=(self.width - width*columnNum - \
                            1*(columnNum-1)) / 2 + counter*(width + 1), y=(self.height - height*rowNumber\
                                - 1*(rowNumber-1)) / 2 + x*(height+1))
                        
                for x in self.mainPanelLabel[view]:# place labels
                    self.currentLabel.append(tkinter.Label(self.__root, text=x, fg="white", bg="black"))

                    font = Font(font=self.currentLabel[-1]["font"])  # get font information
                    lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                    lineWidth = font.measure(x)
                    # print(lineHeight,lineWidth)
                    # print()
#       
                    self.currentLabel[-1].place(x=(self.width - lineWidth) / 2, y= 1)
                    
                buttonTitle = self.mainPanelButtons[view].keys()
                counter = 0
                for x in buttonTitle:# place buttons
                    tem = self.__lambdaCreater(self.mainPanelButtons[view][x][0])
                    # print(self.mainPanelButtons[view][x][0])
                    self.currentButton.append(tkinter.Button(self.__root, text=x, command=tem))

                    font = Font(font=self.currentButton[-1]["font"])  # get font information
                    lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                    lineWidth = font.measure(x)
                    # print(lineHeight,lineWidth)
                    # print()
#       
                    self.currentButton[-1].place(x=(self.width - lineWidth) * (counter * 2 + 1) / (len(buttonTitle)*2),
                        y=self.height - lineHeight*2)
                    counter += 1
#       
            elif mode == "formConfig":
                self.widgetsCleaner(self.currentButton, self.currentLabel, self.currentInput, \
                    self.currentOtherWidgets)

                self.windowSize[lastmode] = []  # back up the size of last window
                self.windowSize[lastmode].append(self.width)
                self.windowSize[lastmode].append(self.height)
                self.windowSize[lastmode].append(self.positionX)
                self.windowSize[lastmode].append(self.positionY)

                self.width = int(self.__root.winfo_screenwidth() / 5)
                self.height = int(self.width * 2 / 1)
                self.positionX = 50
                self.positionY = 50
                
                for x in self.mainPanelLabel[view]:# place buttons
                    counter += 1
                    self.currentLabel.append(tkinter.Label(self.__root, text=x, fg="white", bg="black"))

                    font = Font(font=self.currentLabel[-1]["font"])  # get font information
                    lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                    lineWidth = font.measure(x)
                    # print(lineHeight,lineWidth)
                    # print()
#       
                    self.currentLabel[-1].place(x=(self.width - lineWidth) / 2,
                                                 y=counter * self.height / (buttonNum + 1) - lineHeight / 2)

            elif mode == "binary calculation":
                self.widgetsCleaner(self.currentButton, self.currentLabel, self.currentInput, \
                    self.currentOtherWidgets)

                self.width = int(self.__root.winfo_screenwidth() / 1.25)
                self.height = int(self.__root.winfo_screenheight() / 4)
                self.positionX = 50
                self.positionY = 50
                
                rowNumber = 1
                columnNum = 3
                com = ttk.Combobox(self.__root)     # #创建下拉菜单
                com.place(x = 0, y = 0)     # #将下拉菜单绑定到窗体
                self.__root.update()
                com["value"] = ("test1", "test2", "test3")    # #给下拉菜单设定值
                com.current(2)    # #设定下拉菜单的默认值为第3个
                width = com.winfo_width()
                height = com.winfo_height()
                com.destroy()
                
                titleKeys = list(self.mapForm.keys())
                
                self.currentOtherWidgets.append(ttk.Combobox(self.__root))
                self.currentOtherWidgets[-1]["value"] = titleKeys
                self.currentOtherWidgets[-1].current(0)
                
                self.currentOtherWidgets[-1].place(x=(self.width - width*columnNum - \
                    1*(columnNum-1)) / 2 + 0*(width + 1), y=(self.height - height*rowNumber\
                        - 1*(rowNumber-1)) / 2 + 0*(height+1))

                self.currentOtherWidgets.append(ttk.Combobox(self.__root))
                self.currentOtherWidgets[-1]["value"] = self.binaryCalculation
                self.currentOtherWidgets[-1].current(0)
                
                self.currentOtherWidgets[-1].place(x=(self.width - width*columnNum - \
                    1*(columnNum-1)) / 2 + 1*(width + 1), y=(self.height - height*rowNumber\
                        - 1*(rowNumber-1)) / 2 + 0*(height+1))
                
                self.currentOtherWidgets.append(ttk.Combobox(self.__root))
                self.currentOtherWidgets[-1]["value"] = titleKeys
                self.currentOtherWidgets[-1].current(0)
                
                self.currentOtherWidgets[-1].place(x=(self.width - width*columnNum - \
                    1*(columnNum-1)) / 2 + 2*(width + 1), y=(self.height - height*rowNumber\
                        - 1*(rowNumber-1)) / 2 + 0*(height+1))
                        
                x = "Please decide which binary computation between two column. Column: " + titleKeys[self.formConditionSelection[self.specialData]]
                self.currentLabel.append(tkinter.Label(self.__root, text=x, fg="white", bg="black"))
    
                font = Font(font=self.currentLabel[-1]["font"])  # get font information
                lineWidth = font.measure(x)
#      
                self.currentLabel[-1].place(x=(self.width - lineWidth) / 2, y= 1)
                    
                    
                
                self.currentButton.append(tkinter.Button(self.__root, text="Back", command=lambda: self.Start(6544)))
        
                font = Font(font=self.currentButton[-1]["font"])  # get font information
                lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                lineWidth = font.measure("Back")
                # print(lineHeight,lineWidth)
                # print()
#      
                self.currentButton[-1].place(x=(self.width - lineWidth) * (0 * 2 + 1) / (2*2),
                    y=self.height - lineHeight*2)
                
                self.currentButton.append(tkinter.Button(self.__root, text="Confirm", command=lambda: self.Start(6545)))
        
                font = Font(font=self.currentButton[-1]["font"])  # get font information
                lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                lineWidth = font.measure("Confirm")
                # print(lineHeight,lineWidth)
                # print()
#      
                self.currentButton[-1].place(x=(self.width - lineWidth) * (1 * 2 + 1) / (2*2),
                    y=self.height - lineHeight*2)

            elif mode == "time":
                self.widgetsCleaner(self.currentButton, self.currentLabel, self.currentInput, \
                    self.currentOtherWidgets)

                self.width = int(self.__root.winfo_screenwidth() / 1.25)
                self.height = int(self.__root.winfo_screenheight() / 4)
                self.positionX = 50
                self.positionY = 50
                
                rowNumber = 1
                columnNum = 1
                com = ttk.Combobox(self.__root)     # #创建下拉菜单
                com.place(x = 0, y = 0)     # #将下拉菜单绑定到窗体
                self.__root.update()
                com["value"] = ("test1", "test2", "test3")    # #给下拉菜单设定值
                com.current(2)    # #设定下拉菜单的默认值为第3个
                width = com.winfo_width()
                height = com.winfo_height()
                com.destroy()
                
                titleKeys = list(self.mapForm.keys())
                
                self.currentOtherWidgets.append(ttk.Combobox(self.__root))
                self.currentOtherWidgets[-1]["value"] = self.timeCondition
                self.currentOtherWidgets[-1].current(0)
                
                self.currentOtherWidgets[-1].place(x=(self.width - width*columnNum - \
                    1*(columnNum-1)) / 2 + 0*(width + 1), y=(self.height - height*rowNumber\
                        - 1*(rowNumber-1)) / 2 + 0*(height+1))
                        
                x = "Please decide type of time you would like to record. Column: " + titleKeys[self.formConditionSelection[self.specialData]]
                self.currentLabel.append(tkinter.Label(self.__root, text=x, fg="white", bg="black"))
    
                font = Font(font=self.currentLabel[-1]["font"])  # get font information
                lineWidth = font.measure(x)
#      
                self.currentLabel[-1].place(x=(self.width - lineWidth) / 2, y= 1)
                print(self.width, lineWidth)
                    
                    
                
                self.currentButton.append(tkinter.Button(self.__root, text="Back", command=lambda: self.Start(6544)))
        
                font = Font(font=self.currentButton[-1]["font"])  # get font information
                lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                lineWidth = font.measure("Back")
                # print(lineHeight,lineWidth)
                # print()
#      
                self.currentButton[-1].place(x=(self.width - lineWidth) * (0 * 2 + 1) / (2*2),
                    y=self.height - lineHeight*2)
                
                self.currentButton.append(tkinter.Button(self.__root, text="Confirm", command=lambda: self.Start(6545)))
        
                font = Font(font=self.currentButton[-1]["font"])  # get font information
                lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                lineWidth = font.measure("Confirm")
                # print(lineHeight,lineWidth)
                # print()
#      
                self.currentButton[-1].place(x=(self.width - lineWidth) * (1 * 2 + 1) / (2*2),
                    y=self.height - lineHeight*2)

            elif mode == "special condition":
                self.widgetsCleaner(self.currentButton, self.currentLabel, self.currentInput, \
                    self.currentOtherWidgets)

                self.width = int(self.__root.winfo_screenwidth() / 1.25)
                self.height = int(self.__root.winfo_screenheight() / 4)
                self.positionX = 50
                self.positionY = 50
                
                rowNumber = 1
                columnNum = 5
                com = ttk.Combobox(self.__root)     # #创建下拉菜单
                com.place(x = 0, y = 0)     # #将下拉菜单绑定到窗体
                self.__root.update()
                com["value"] = ("test1", "test2", "test3")    # #给下拉菜单设定值
                com.current(2)    # #设定下拉菜单的默认值为第3个
                width = com.winfo_width()
                height = com.winfo_height()
                com.destroy()
                
                titleKeys = list(self.mapForm.keys())
                
                self.currentOtherWidgets.append(ttk.Combobox(self.__root))
                self.currentOtherWidgets[-1]["value"] = ["If..."] + titleKeys
                self.currentOtherWidgets[-1].current(0)
                
                self.currentOtherWidgets[-1].place(x=(self.width - width*columnNum - \
                    1*(columnNum-1)) / 2 + 0*(width + 1), y=(self.height - height*rowNumber\
                        - 1*(rowNumber-1)) / 2 + 0*(height+1))

                self.currentOtherWidgets.append(ttk.Combobox(self.__root))
                self.currentOtherWidgets[-1]["value"] = self.specialCondition
                self.currentOtherWidgets[-1].current(0)
                
                self.currentOtherWidgets[-1].place(x=(self.width - width*columnNum - \
                    1*(columnNum-1)) / 2 + 1*(width + 1), y=(self.height - height*rowNumber\
                        - 1*(rowNumber-1)) / 2 + 0*(height+1))
                
                self.currentOtherWidgets.append(ttk.Combobox(self.__root))
                self.currentOtherWidgets[-1]["value"] = titleKeys
                
                self.currentOtherWidgets[-1].place(x=(self.width - width*columnNum - \
                    1*(columnNum-1)) / 2 + 2*(width + 1), y=(self.height - height*rowNumber\
                        - 1*(rowNumber-1)) / 2 + 0*(height+1))
                
                self.currentOtherWidgets.append(ttk.Combobox(self.__root))
                self.currentOtherWidgets[-1]["value"] = ["Then..."] + [self.formDataDesicion[1]] + ["Do nothing"]
                self.currentOtherWidgets[-1].current(0)
                
                self.currentOtherWidgets[-1].place(x=(self.width - width*columnNum - \
                    1*(columnNum-1)) / 2 + 3*(width + 1), y=(self.height - height*rowNumber\
                        - 1*(rowNumber-1)) / 2 + 0*(height+1))
                
                
                self.currentOtherWidgets.append(ttk.Combobox(self.__root))
                self.currentOtherWidgets[-1]["value"] = ["Else..."] + [self.formDataDesicion[1]] + ["Do nothing"]
                self.currentOtherWidgets[-1].current(0)
                
                self.currentOtherWidgets[-1].place(x=(self.width - width*columnNum - \
                    1*(columnNum-1)) / 2 + 4*(width + 1), y=(self.height - height*rowNumber\
                        - 1*(rowNumber-1)) / 2 + 0*(height+1))
                        
                x = "Please decide condition between two column and the operation when \
condition is met.(or a number and a column). Column: " + titleKeys[self.formConditionSelection[self.specialData]]
                self.currentLabel.append(tkinter.Label(self.__root, text=x, fg="white", bg="black"))
    
                font = Font(font=self.currentLabel[-1]["font"])  # get font information
                lineWidth = font.measure(x)
#      
                self.currentLabel[-1].place(x=(self.width - lineWidth) / 2, y= 1)
                    
                    
                
                self.currentButton.append(tkinter.Button(self.__root, text="Back", command=lambda: self.Start(6544)))
        
                font = Font(font=self.currentButton[-1]["font"])  # get font information
                lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                lineWidth = font.measure("Back")
                # print(lineHeight,lineWidth)
                # print()
#      
                self.currentButton[-1].place(x=(self.width - lineWidth) * (0 * 2 + 1) / (2*2),
                    y=self.height - lineHeight*2)
                
                self.currentButton.append(tkinter.Button(self.__root, text="Confirm", command=lambda: self.Start(6545)))
        
                font = Font(font=self.currentButton[-1]["font"])  # get font information
                lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                lineWidth = font.measure("Confirm")
#      
                self.currentButton[-1].place(x=(self.width - lineWidth) * (1 * 2 + 1) / (2*2),
                    y=self.height - lineHeight*2)
                

            self.__root.geometry("{0}x{1}+{2}+{3}" \
                .format(self.width, self.height, self.positionX, self.positionY))

            if align == "downward":
                for x in self.mainPanelButtons[view].keys():# place buttons
                    counter += 1
                    tem = self.__lambdaCreater(self.mainPanelButtons[view][x][0])
                    # print(self.mainPanelButtons[view][x][0])
                    self.currentButton.append(tkinter.Button(self.__root, text=x, command=tem))
                    if len(buttonStatusSetup) > 0:
                        # print(counter)
                        # print(buttonStatusSetup)
                        # print(len(self.mainPanelInput[view]))
                        # print(buttonStatusSetup[counter - len(self.mainPanelInput[view]) - 1])
                        self.currentButton[-1].config(state=buttonStatus[buttonStatusSetup[counter - len(self.mainPanelInput[view]) - 1]])

                    font = Font(font=self.currentButton[-1]["font"])  # get font information
                    lineHeight = font.metrics("linespace")  # calculate hieght and weidth by font information
                    lineWidth = font.measure(x)
                    # print(lineHeight,lineWidth)
                    # print()
#       
                    self.currentButton[-1].place(x=(self.width - lineWidth) / 2,
                                                 y=counter * self.height / (buttonNum + 1) - lineHeight / 2)



    def __lambdaCreater(self, x):  # create lambda function for button, prevent shollow copy
        return lambda: self.Start(x)

    def subWindowCreater(self, width, height, lastMode = "root", x = 100, y = 100, \
        listener = "None", alphaValue = 0.8, bgColor = "black", scrollBar = False, backup = True):
        
        self.__subWindows = tkinter.Toplevel()  # set up sub window
        self.__subWindows.title("recorded operations")
        self.__subWindows.iconbitmap("agent.ico")   # 更改窗口图标
        if backup:
            self.windowSize[lastMode] = []  # back up the size of last window
            self.windowSize[lastMode].append(self.width)
            self.windowSize[lastMode].append(self.height)
            self.windowSize[lastMode].append(self.positionX)
            self.windowSize[lastMode].append(self.positionY)
        
        self.width = width
        self.height = height
        self.positionX = x
        self.positionY = y

        self.__subWindows.attributes("-alpha", alphaValue)

        self.__subWindows.geometry("{0}x{1}+{2}+{3}" \
                                   .format(self.width, self.height, x, y))
        self.__subWindows.configure(bg=bgColor)

        if scrollBar:
            self.scroller = ttk.Scrollbar(self.__subWindows)  #设置窗口滚动条
            self.scroller.pack(side = RIGHT, fill = Y)  #设置窗口滚动条位置
            
            self.mylistBox = tkinter.Listbox(self.__subWindows, yscrollcommand = self.scroller.set, \
                width=self.width, bg = bgColor, fg = "white")  #创建列表框

            
            self.mylistBox.pack( side = LEFT , fill="both")  
            self.scroller.config( command = self.mylistBox.yview )  
        
        if listener == "mouse" and self.listener != None:
            self.listener = eventMouse()
            self.listener.StartListener()

        self.__subWindows.resizable(0, 0)

    def screenShotCreation(self, alphaValue=0.5, bgColor="black") -> None:
        """_summary_

        Args:
            alphaValue (int): the visibility of the screen shot window
            bgColor (str): allow for user to custom the back ground color
        """
        self.screenShotor = tkinter.Toplevel()  # set up sub window
        self.windowSize["root"] = []  # back up the size of root window
        self.windowSize["root"].append(self.width)
        self.windowSize["root"].append(self.height)
        self.windowSize["root"].append(self.positionX)
        self.windowSize["root"].append(self.positionY)
        
        self.width = self.screenShotor.winfo_screenwidth()  # make it large as the screen
        self.height = self.screenShotor.winfo_screenheight()

        self.screenShotor.overrideredirect(True)  # remove tk default component(e.g. window close button)
        self.screenShotor.attributes("-alpha", alphaValue)

        self.screenShotor.geometry("{0}x{1}+{2}+{3}" \
                                   .format(self.width, self.height, 0, 0))
        self.screenShotor.configure(bg=bgColor)
        self.__num = 6
        self.canvasPlace(target="sub")

        self.listener.restart()
        self.screenShot = 1
        print("Screenshoter setted up")
        if self.listener == None:
            self.listener = eventMouse()
            self.listener.StartListener()

    def Start(self, status):
        print("111111111111111111111111")
        print(status)
        # print("Start:",self.statusID)
        if status == 1100:  #
            self.loopCounter = ([],[],[],[])
            self.recordPointer = len(self.userInteraction.recoredBehaviours)
            print(self.recordPointer)
            print(self.userInteraction.recoredBehaviours)
            self.layOutController("record", 1, "root")
            if self.__subWindows == None:
                self.subWindowInfo = (int(self.__root.winfo_screenwidth() / 6), int(self.width * 2 / 1),\
                    int(self.__root.winfo_screenwidth() - self.__root.winfo_screenwidth() / 5), 50)
                self.subWindowCreater(self.subWindowInfo[0], self.subWindowInfo[1], \
                    x = self.subWindowInfo[2], y = self.subWindowInfo[3], scrollBar = True, backup = False)
            self.statusID = 1010
            
        elif status == 1101:
            # print(self.mainPanelButtons)
            self.layOutController("executeList", 2, "root")
            self.statusID = 4000
            
        elif status == 1104:
            self.screenShotCreation()
            # self.subWindows.append(windowsUI(True,0.5,"black",listener=mouseL,screenShot=3))
            # print("!!!!!!!!!!!!!!!!!!")
            self.__root.attributes("-alpha", 0)
            if self.__subWindows != None:
                self.__subWindows.attributes("-alpha", 0)
            self.statusID = -1000
            # print("Start:",self.statusID)
            
        elif status == 1110:
            self.layOutController("clickRecord", 3, "record")
            self.statusID = 1020
            
        elif status == 1114:
            self.layOutController("textRecognizeConfig", 9, "record")
            self.statusID = 6010
            
        elif status == 1115:
            self.layOutController("loopConfig", 8, "record")
            self.statusID = 6000
            
        elif status == 1116:
            self.layOutController("standbyConfig", 5, "record")
            self.statusID = 5000
            
        elif status == 1117:
            self.layOutController("formConfig", 17, "root")
            self.statusID = 6510
            
        elif status == 1118:
            self.recordPointer = len(self.userInteraction.recoredBehaviours)
            self.userInteraction.oldPointer = len(self.userInteraction.recoredBehaviours)
            if len(self.mainPanelButtons[2].keys()) == 1: # add button to execute panel
                self.mainPanelButtons[2]["1"] = [4101]
            else:
                self.mainPanelButtons[2][str(int(list(self.mainPanelButtons[2].keys())[-1]) + 1)] = \
                    [self.mainPanelButtons[2][list(self.mainPanelButtons[2].keys())[-1]][0] + 1]
            print(self.mainPanelButtons)
                
            self.layOutController(lastmode = "record")
            self.messageBox = showinfo("RseMessager", "operation recorded!")
            self.statusID = 1000
            
        elif status == 1119:
            self.userInteraction.delRecord()
            self.recordPointer = len(self.userInteraction.recoredBehaviours)
            
            self.layOutController(lastmode = "record")
            self.statusID = 1000
            
        elif status == 1120:
            self.messageBox = showinfo("RseMessager", "Please click what you would like us to\
click after pressing 'ok'. This message could be close in setting panel.")
            self.__root.attributes("-alpha", 0)
            if self.__subWindows != None:
                self.__subWindows.attributes("-alpha", 0)
            self.listener.restart()
            self.statusID = 2010
            
        elif status == 1121:
            fileList = self.dbManagement.UIImageList()
            # print(fileList)
            self.mainPanelButtons[4].clear()
            self.mainPanelButtons[4]["back"] = [1130]
            for x in range(len(fileList)):
                self.mainPanelButtons[4]["button. "+fileList[x]] = [1130 + x +1]
            self.layOutController("UIClick", 4, "clickRecord")
            self.statusID = 1030
            
        elif status == 1122:
            self.layOutController("record", 1, "clickRecord")
            self.statusID = 1010
            
        elif status == 1130:
            self.layOutController("clickRecord", 3, "UIClick")
            self.statusID = 1020
            
        elif 2000>status >= 1131:
            fileName = list(self.mainPanelButtons[4].keys())[status % 1130][8:]
            image = cv2.imread(self.dbManagement.currentPath + "\\UI\\" + fileName, flags=1)
            print(self.dbManagement.currentPath + "\\UI\\" + fileName)
            self.userInteraction.actionRecord(self.description[1131],\
                (image, (0,0,self.screen.width, self.screen.height)), self.recordPointer)
            self.shownListManagement("add", self.description[1131]+": " + fileName)
            
        elif status == 4100:
            self.layOutController(lastmode = "record")
            self.statusID = 1000
            
        elif 4200> status >= 4101:
            del self.temForm
            self.temForm = dict()
            self.currentColumnNum = -1
            self.counter = -1
            self.executePointer = 0
            self.loopCounter = ([],[],[],[])
            self.statusID = 4200 + (status % 4100)
            
        elif status == 5100:
            self.layOutController("timeWaitConfig", 6, "standbyConfig")
            self.statusID = 5010
            
        elif status == 5102:
            fileList = self.dbManagement.UIImageList()
            # print(fileList)
            self.mainPanelButtons[7].clear()
            self.mainPanelButtons[7]["back"] = [5120]
            for x in range(len(fileList)):
                self.mainPanelButtons[7]["button. "+fileList[x]] = [5120 + x +1]
            self.layOutController("UIClick", 7, "standbyConfig")
            self.statusID = 5020
            
        elif status == 5103:
            self.layOutController("record", 1, "standbyConfig")
            self.statusID = 1010
            
        elif status == 5110:
            counter = 0
            for x in range(len(self.currentInput)):
                tem = self.currentInput[x].get()
                if tem.isdigit():
                    counter += int(tem)*60**(2-x)
            self.userInteraction.actionRecord(self.description[status],\
                (counter,), self.recordPointer)
            self.shownListManagement("add", self.description[status]+": " + str(counter) + "s")
            self.layOutController("standbyConfig", 5, "timeWaitConfig")
            self.statusID = 5000
            
        elif status == 5111:
            self.layOutController("standbyConfig", 5, "timeWaitConfig")
            self.statusID = 5000
            
        elif status == 5120:
            self.layOutController("standbyConfig", 5, "UIClick")
            self.statusID = 5000
            
        elif 6000 > status >=5121:
            fileName = list(self.mainPanelButtons[7].keys())[status % 5120][8:]
            image = cv2.imread(self.dbManagement.currentPath + "\\UI\\" + fileName, flags=1)
            self.userInteraction.actionRecord(self.description[5121],\
                (image, (0,0,self.screen.width, self.screen.height)), self.recordPointer)
            self.shownListManagement("add", self.description[5121]+": " + fileName)
        
        elif status == 6100:
            self.layOutController("record", 1, "loopConfig")
            self.statusID = 1010
        
        elif 6104 >= status >=6101:
            if len(self.loopCounter[status % 6101]) < 3:
                tem = self.currentInput[0].get()
                if tem.isdigit() or tem == "-1":
                    counter = int(tem)

                    self.loopCounter[status % 6101].clear()
                    if self.recordPointer >= len(self.userInteraction.recoredBehaviours):
                        self.loopCounter[status % 6101].append(0)
                    else:
                        self.loopCounter[status % 6101].append(len(self.userInteraction.recoredBehaviours\
                            [self.recordPointer]))
                        
                    self.loopCounter[status % 6101].append(counter)
                    self.loopCounter[status % 6101].append(False)

                    self.shownListManagement("add", "Loop " + chr(65 + status % 6101)+ " Start")
                    self.layOutController("record", 1, "loopConfig")
                    self.statusID = 1010
                else:
                    self.messageBox = showinfo("RseMessager", "Please provide an integer")
            else:
                self.loopCounter[status % 6101][2] = True
                
                self.userInteraction.actionRecord(self.description[6101],\
                    (self.loopCounter[status % 6101][0], self.loopCounter[status % 6101][1], \
                        status % 6101), self.recordPointer)
                self.shownListManagement("add", "Loop " + chr(65 + status % 6101)+ " Stop at "\
                    + str(self.loopCounter[status % 6101][1]) + " times")
                self.layOutController("record", 1, "loopConfig")
                self.statusID = 1010
        
        elif status == 6111:
            self.messageBox = askquestion("RseMessager", "We need to set this up by following 3 steps:\n\
1. choose or build up a model to find text boxes.\n2. choose or build up a model to make sure target\
text are included by recognization area.\n3. test the result.\n Would you like to continue?")
            print(self.messageBox)
            if self.messageBox == "yes":
                for x in range(len(list(self.ocr.modelID))):
                    self.mainPanelButtons[10][str(self.ocr.modelID[x])] = [self.mainPanelButtons[10]\
                        [list(self.mainPanelButtons[10].keys())[-1]][0] + 1]

                self.layOutController("textBoxesModelling", 10, "record")
                self.statusID = 6020
                
        elif 6200 > status >=6122:
            self.testModelID = status % 6122
            self.messageBox = showinfo("RseMessager", "Text boxes recognize model has been ready,\
we would need to test this model worked with your circumstance. Please tell us \
to area of the list after pressing 'ok'.(Press key 'alt+z' finish the step)")
            self.screenShotCreation()
            # self.subWindows.append(windowsUI(True,0.5,"black",listener=mouseL,screenShot=3))
            # print("!!!!!!!!!!!!!!!!!!")
            self.__root.attributes("-alpha", 0)
            if self.__subWindows != None:
                self.__subWindows.attributes("-alpha", 0)
            self.statusID = 2000
        
        elif status == 6200:
            for x in range(len(list(self.ocr.modelID))):
                self.mainPanelButtons[11][str(self.ocr.modelID[x])] = [self.mainPanelButtons[11]\
                    [list(self.mainPanelButtons[11].keys())[-1]][0] + 1]
            self.layOutController("textSeeker", 11, "textBoxesModelling")
            self.statusID = 6201
            
        elif status == 6202:
            fileList = self.dbManagement.UIImageList()
            # print(fileList)
            self.mainPanelButtons[12].clear()
            self.mainPanelButtons[12]["No"] = [6205]
            for x in range(len(fileList)):
                self.mainPanelButtons[12]["button. "+fileList[x]] = [6205 + x +1]
            self.layOutController("iconFilterOut", 12, "textSeeker")
            self.statusID = 6203
            
        elif status == 6205:
            self.messageBox = showinfo("RseMessager", "In the next step, the The model would try to \
find out all target text, please tell us the target text of the first line.(Press key 'alt+z' finish the step)")
            self.__root.attributes("-alpha", 0)
            if self.__subWindows != None:
                self.__subWindows.attributes("-alpha", 0)
            self.__root.geometry("{0}x{1}+{2}+{3}" \
                .format(self.width, self.height, 0-self.width, 0-self.height))
            self.__subWindows.geometry("{0}x{1}+{2}+{3}" \
                .format(self.subWindowInfo[0], self.subWindowInfo[1], 0 - self.subWindowInfo[0], 0 - self.subWindowInfo[1]))
                
            self.screenShotCreation()
            self.counter = 0
            self.statusID = 2002
            
        elif 6300 > status > 6205:
            fileName = list(self.mainPanelButtons[12].keys())[status % 6205][8:]
            print(self.dbManagement.currentPath + "\\UI\\" + fileName)
            self.imageFilterOut = cv2.imread(self.dbManagement.currentPath + "\\UI\\" + fileName, flags=1)
            self.layOutController("positionSelect", 13, "iconFilterOut")
            self.statusID = 6204
            
        elif 6400 > status >= 6302:
            self.testModelID = status % 6302
            self.messageBox = showinfo("RseMessager", "Error location filtering model has been ready,\
we would need to test this model worked with your circumstance. Please tell us arbitrary area \
the model would determine if this cut any text(you could test multiple area at the same time). \
(Press key 'alt+z' finish the step)")
            self.screenShotCreation()
            # self.subWindows.append(windowsUI(True,0.5,"black",listener=mouseL,screenShot=3))
            # print("!!!!!!!!!!!!!!!!!!")
            self.__root.attributes("-alpha", 0)
            if self.__subWindows != None:
                self.__subWindows.attributes("-alpha", 0)
            self.statusID = 2001
            
        elif status == 6400 or status == 6401:
            self.indexFilterOut = -1 * (status % 6400)
            print(self.indexFilterOut)
            self.messageBox = showinfo("RseMessager", "In the next step, the model would try to \
find out all target text, please tell us the target text of the first line.(Press key 'alt+z' finish the step)")
            self.__root.attributes("-alpha", 0)
            if self.__subWindows != None:
                self.__subWindows.attributes("-alpha", 0)
                
            self.__root.geometry("{0}x{1}+{2}+{3}" \
                .format(self.width, self.height, 0-self.width, 0-self.height))
            self.__subWindows.geometry("{0}x{1}+{2}+{3}" \
                .format(self.subWindowInfo[0], self.subWindowInfo[1], 0 - self.subWindowInfo[0], 0 - self.subWindowInfo[1]))
            

            self.screenShotCreation()
            self.counter = 0
            self.statusID = 2002

        elif status == 6402:
            self.messageBox = askquestion("RseMessager", "Is there any text not on form\
? If so, we need extra steps to acquire the text.(Please also include the step to turn back to the form.)")
            if self.messageBox == "yes":
                self.layOutController("extraSteps", 14, "positionSelect")
                self.statusID = 6403
            else:
                self.layOutController("record", 1, "positionSelect")
                self.statusID = 1010

        elif status == 6406:
            self.layOutController("extraSteps", 14, "TextRetrievalMode")
            self.statusID = 6403

        elif status == 6410:
            textBoxesParam = (self.description[2010], "textBoxCentre")
            self.sequencialCommand.append(copy.deepcopy(textBoxesParam))
            self.messageBox = showinfo("RseMessager", "step recorded")
            self.statusID = 6403
            
        elif status == 6411:
            fileList = self.dbManagement.UIImageList()
            for x in range(len(fileList)):
                self.mainPanelButtons[15]["button. "+fileList[x]] = [6420 + x]
            self.layOutController("extraStepsUIClick", 15, "extraSteps")
            self.statusID = 6404
            
        elif status == 6412:
            self.layOutController("TextRetrievalMode", 16, "extraSteps")
            self.statusID = 6405
            
        elif status == 6413:
            self.layOutController("record", 1, "positionSelect")
            self.statusID = 1010
            
        elif 6500 > status >= 6420:
            fileName = list(self.mainPanelButtons[15].keys())[status % 6420][8:]
            image = cv2.imread(self.dbManagement.currentPath + "\\UI\\" + fileName, flags=1)
            
            textBoxesParam = (self.description[1131],\
                (image, (0,0,self.screen.width, self.screen.height)))
            self.sequencialCommand.append(copy.deepcopy(textBoxesParam))
            self.messageBox = showinfo("RseMessager", "step recorded")
            self.layOutController("extraSteps", 14, "extraStepsUIClick")
            self.statusID = 6403
            
        elif status == 6500:
            self.messageBox = showinfo("RseMessager", "Please tell us the location of the text.")
            self.__root.attributes("-alpha", 0)
            if self.__subWindows != None:
                self.__subWindows.attributes("-alpha", 0)
                
            self.screenShotCreation()
            self.counter = 0
            self.statusID = 2004
        
        elif status == 6501:
            self.messageBox = showinfo("RseMessager", "Please tell us the location of the text.")
            self.__root.attributes("-alpha", 0)
            if self.__subWindows != None:
                self.__subWindows.attributes("-alpha", 0)
                
            self.screenShotCreation()
            self.counter = 0
            self.statusID = 2003

        elif status == 6520:
            pass
        
        elif status == 6521:
            self.temFileName = "人员信息统计.xlsx"
            tem = []
            for x in range(len(list(self.temForm.keys()))):
                if x == 0:
                    tem.append(len(list(self.temForm.keys())))
                tem.append(self.temForm[x][0])
            if len(tem) > 0:
                self.mainPanelCombobox[18].append(tem)
            
            self.layOutController("memberInfoForm", 18, "formConfig", align = "form")
            self.statusID = 6530
            
        elif status == 6522:
            pass

        elif status == 6541:
            title = dict()
            firstRow = []
            nonDataRecord = []
            for x in range(len(self.currentOtherWidgets)):
                if x < self.mainPanelCombobox[18][0][0]:
                    title[self.currentOtherWidgets[x].get()] = -1
                else:
                    name = list(title.keys())
                    firstRow.append(self.currentOtherWidgets[x].get())
                    for y in self.temForm.keys():
                        if firstRow[-1] == self.temForm[y][0]:
                            title[name[x - self.mainPanelCombobox[18][0][0]]] = copy.deepcopy(y)
                    if title[name[x - self.mainPanelCombobox[18][0][0]]] == -1:
                        title[name[x - self.mainPanelCombobox[18][0][0]]] = firstRow[-1]
                        nonDataRecord.append(x - self.mainPanelCombobox[18][0][0])
            print(title)
            print(nonDataRecord)
            
            self.formConditionSelection = nonDataRecord
            self.mapForm = title
            tem = copy.deepcopy(self.mainPanelCombobox[18][0])
            self.mainPanelCombobox[18].clear()
            self.mainPanelCombobox[18].append(tem)
            
            self.Start(6542)
            
        elif status == 6542:
            if len(self.formConditionSelection) > self.specialData:
                titleKeys = list(self.mapForm.keys())
                condition = self.mapForm[titleKeys[self.formConditionSelection[self.specialData]]]
                if condition == "--==//binary operation--==//":
                    self.layOutController("binary calculation", align = "form")
                elif condition == "--==//time--==//":
                    self.layOutController("time", align = "form")
                elif condition == "--==//special condition--==//":
                    self.layOutController("special condition", align = "form")
                self.specialData += 1
                self.statusID = 6543
                
            else:
                pass
            
        elif status == 6544:
            pass
        
        elif status == 6545:
            tem = []
            for x in range(len(self.currentOtherWidgets)):
                tem.append(self.currentOtherWidgets[x].get())
                
            titleKeys = list(self.mapForm.keys())
            self.mapForm[titleKeys[self.formConditionSelection[self.specialData - 1]]] = [self.mapForm[titleKeys[self.formConditionSelection[self.specialData - 1]]]] + tem
            
            print(self.specialData)
            print(self.mapForm)
            print(self.formConditionSelection)
            if len(self.formConditionSelection) > self.specialData:
                condition = self.mapForm[titleKeys[self.formConditionSelection[self.specialData]]]
                print(condition)

                if condition == "--==//binary operation--==//":
                    self.layOutController("binary calculation", align = "form")
                elif condition == "--==//time--==//":
                    self.layOutController("time", align = "form")
                elif condition == "--==//special condition--==//":
                    self.layOutController("special condition", align = "form")
                    
                self.specialData += 1
                self.statusID = 6543
            else:
                print(self.mapForm)
                self.userInteraction.actionRecord(self.description[6543],\
                    (self.temFileName, list(self.mapForm.values()), [], [], list(self.mapForm.keys()), \
                        copy.deepcopy(self.formConditionSelection)), self.recordPointer)
                self.shownListManagement("add", self.description[6543]+": " + self.temFileName)
                
                self.specialData = 0
                self.layOutController("record", 1, "special condition")
                self.statusID = 1010

        elif status == 7100:
            self.messageBox = showinfo("RseMessager", "Almost done! We now need to adjust \
scorlling program by 3 test. Please press key 'alt+z' when the following line of data approach to the \
top of scrolling area(Please do not make its text box attached to the top of scrolling area).\n \
    Data:\n \
        " + str(self.landMark))
            
            self.counter = 2
            self.statusID = 7001

    def keeper(self) -> None:
        # print("ID:",self.screenShot)
        print("Status:", self.statusID)
        self.keyBoardInterrupt.activeFlagSet(1)
        self.listener.activeFlagSet(1)
        self.eventAction() # check any action need to take
        
        if self.screenShot == 1:
            if self.drawer() == 1:
                self.__root.after(100, self.keeper)

        else:  # mian panel mode
            self.__root.after(100, self.keeper)

        self.tester()

            
    def tester(self):
        if self.statusID == -1000:
            if self.keyBoardInterrupt.statusGet() == 2:
                # print("ssssssssssssssss",self.statusID)
                self.x = -10
                self.y = -10
                self.screenShot = 2

                for x in self.__rec:
                    transformed = self.transform(x)
                    # if self.counter == 0:
                    #     self.textAreUpperBound = min(transformed["top"] , self.textAreUpperBound)
                    #     self.textAreLowerBound = max(transformed["top"] + transformed["height"],\
                    #         self.textAreLowerBound)
                    # print(self.textAreUpperBound, self.textAreLowerBound)
                    
                    self.recoredArea.append(transformed)
                    print(transformed)
                self.widgetsCleaner(self.__canvas, self.screenShotor)
                self.__rec.clear()
                self.width, self.height, self.positionX, self.positionY = self.windowSize["root"]
                
                self.statusID = -1
        elif self.statusID == -1010:
            if self.keyBoardInterrupt.statusGet() == 2:
                self.x = -10
                self.y = -10
                self.screenShot = 2
                self.widgetsCleaner(self.__canvas, self.screenShotor)
                self.__rec.clear()
                self.recoredArea.clear()
                self.__root.attributes("-alpha", self.alpha)
                if self.__subWindows != None:
                    self.__subWindows.attributes("-alpha", self.alpha)
                self.statusID = 1000
            
        if len(self.recoredArea) > 0 and self.testNum == 0 and self.test: # test
            tem = self.recoredArea.pop()
            self.testScrollingArea = copy.deepcopy(tem)
            # text=self.ocr.areTextTransfer(tem)
            # textBoxes = self.ocr.textboxSeekerTrainer(tem, 50)
            self.textBoxes = self.ocr.textboxSeekerPredictor(tem, 3, 0, True)
            self.statusID = -1010
            self.screenShotCreation()
            counter = 0
            self.testNum = 2
            for x in self.textBoxes:
                leftX = (x["left"] * self.width) / self.screen.width  # 转换相对坐标。
                leftY = (x["top"] * self.height) / self.screen.height

                rightX = ((x["left"] + x["width"]) * self.width) / self.screen.width  # 转换相对坐标。
                rightY = ((x["top"] + x["height"]) * self.height) / self.screen.height
                if counter == 0:
                    self.rectangleCreation(leftX, leftY, rightX, rightY, "crimson")
                elif counter == 1:
                    self.rectangleCreation(leftX, leftY, rightX, rightY, "blue")
                else:
                    self.rectangleCreation(leftX, leftY, rightX, rightY, "green")
                counter += 1
                if counter % 3 == 0:
                    counter = 0
            print(self.recoredArea)
            self.recoredArea.clear()
            # self.dbManagement.OCRModelDataSaver(self.ocr.centroids,\
            #     self.ocr.data, self.ocr.modelID, self.ocr.relativeDistance)
            
        elif len(self.recoredArea) > 0 and self.testNum == 1 and self.test: # test
            
            tem = copy.deepcopy(self.recoredArea)
            # text=self.ocr.areTextTransfer(tem)
            # textBoxes = self.ocr.textboxSeekerTrainer(tem, 50)
            # textBoxes = self.ocr.textboxSeekerPredictor(tem, 50, 0)
            self.ocr.textLocationTrainer(tem, {"top": 0, "left": 0, "width": \
                self.screen.width, "height": self.screen.height})
            
            # self.ocr.textLocationPredictor(tem, {"top": 0, "left": 0, "width": \
            #     self.screen.width, "height": self.screen.height}, 1)
            
            self.testNum = 100
            self.recoredArea.clear()
            # self.statusID = 3000
            # self.screenShotCreation()
            # counter = 0
            # for x in textBoxes:
            #     leftX = (x["left"] * self.width) / self.screen.width  # 转换相对坐标。
            #     leftY = (x["top"] * self.height) / self.screen.height
                
            #     rightX = ((x["left"] + x["width"]) * self.width) / self.screen.width  # 转换相对坐标。
            #     rightY = ((x["top"] + x["height"]) * self.height) / self.screen.height
            #     if counter == 0:
            #         self.rectangleCreation(leftX, leftY, rightX, rightY, "crimson")
            #     elif counter == 1:
            #         self.rectangleCreation(leftX, leftY, rightX, rightY, "blue")
            #     else:
            #         self.rectangleCreation(leftX, leftY, rightX, rightY, "green")
            #     counter += 1
            #     if counter % 3 == 0:
            #         counter = 0
            # print(self.recoredArea)
            # self.dbManagement.OCRModelDataSaver(self.ocr.centroids,\
            #     self.ocr.data, self.ocr.modelID, self.ocr.relativeDistance)
            
        elif len(self.recoredArea) > 0 and self.testNum == 100 and self.test:
            
            tem = copy.deepcopy(self.recoredArea)
            # text=self.ocr.areTextTransfer(tem)
            # textBoxes = self.ocr.textboxSeekerTrainer(tem, 50)
            # textBoxes = self.ocr.textboxSeekerPredictor(tem, 50, 0)
            # self.ocr.textLocationTrainer(tem, {"top": 0, "left": 0, "width": \
            #     self.screen.width, "height": self.screen.height})
            
            self.ocr.textLocationPredictor(tem, {"top": 0, "left": 0, "width": \
                self.screen.width, "height": self.screen.height}, 1)
            
            self.testNum = 100
            self.recoredArea.clear()
            
        elif len(self.recoredArea) > 0 and self.testNum == 2 and self.test:
            result = []
            if len(self.textBoxes) >= 2:
                firstBox = 1
                lastBox = -1
                print()
                if len(self.temForm.keys()) == 0 or self.textBoxes[0]["top"] - self.testScrollingArea["top"] >= \
                    self.testScrollingArea["height"] / 10:
                    
                    firstBox = 0
                if self.testScrollingArea["top"] + self.testScrollingArea["height"] - \
                    (self.textBoxes[-1]["top"] + self.textBoxes[-1]["height"]) >= \
                    self.testScrollingArea["height"] / 10:
                    
                    lastBox = len(self.textBoxes)
                    
                for x in self.textBoxes[firstBox:lastBox]:
                    tem = copy.deepcopy(self.recoredArea)
                    if len(self.ocr.relativeDistance[1]) == 0:
                        print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
                        result += self.ocr.textSeeker(tem, copy.deepcopy(x), {"top": 0, "left": 0, "width": \
                            self.screen.width, "height": self.screen.height}, 1, "train", \
                                self.textAreLowerBound - self.textAreUpperBound, self.textAreUpperBound, \
                                    self.textAreLowerBound)
                    else:
                        print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
                        result += self.ocr.textSeeker(tem, copy.deepcopy(x), {"top": 0, "left": 0, "width": \
                            self.screen.width, "height": self.screen.height}, 1, "predict")
                textMatched = self.ocr.areTextTransfer(self.testScrollingArea, result, True, \
                        self.imageFilterOut, self.indexFilterOut, (0,0,self.screen.width,self.screen.height))
                # text=self.ocr.areTextTransfer(tem)
                # textBoxes = self.ocr.textboxSeekerTrainer(tem, 50)
                # textBoxes = self.ocr.textboxSeekerPredictor(tem, 50, 0)
                # self.ocr.textLocationTrainer(tem, {"top": 0, "left": 0, "width": \
                #     self.screen.width, "height": self.screen.height})
                self.screenShotCreation()
                counter = 0
                print(result)
                formCounter = 0
                for x in range(len(result)):
                    leftX = (result[x]["left"] * self.width) / self.screen.width  # 转换相对坐标。
                    leftY = (result[x]["top"] * self.height) / self.screen.height

                    rightX = ((result[x]["left"] + result[x]["width"]) * self.width) / self.screen.width  # 转换相对坐标。
                    rightY = ((result[x]["top"] + result[x]["height"]) * self.height) / self.screen.height
                    if counter == 0:
                        self.rectangleCreation(leftX, leftY, rightX, rightY, "crimson")
                    elif counter == 1:
                        self.rectangleCreation(leftX, leftY, rightX, rightY, "blue")
                    else:
                        self.rectangleCreation(leftX, leftY, rightX, rightY, "green")
                    counter += 1
                    if counter % 3 == 0:
                        counter = 0
                    if formCounter not in self.temForm.keys():
                        self.temForm[formCounter] = list()

                    self.temForm[formCounter].append(textMatched[x])
                    counter += 1
                    formCounter += 1
                    if counter % 3 == 0:
                        counter = 0
                    if formCounter % len(self.recoredArea) == 0:
                        formCounter = 0


                self.testNum = 4
                if len(self.temForm.keys()) > 0:
                    tem = []
                    for x in range(len(self.temForm.keys())):
                        tem.append(self.temForm[x][-1])
                self.landMark = tem
                self.landMarkPosition = self.textBoxes[lastBox-1]
                self.recoredArea.clear()
                self.statusID = -1000
            
        elif len(self.recoredArea) > 0 and self.testNum == 3 and self.test: # test
            print("----------------------")
            tem = self.recoredArea.pop()
            text = self.ocr.areTextTransfer(tem, textPosition= [{"top": 1058, "left": 1771, \
                "width": 78, "height": 20}], match = True)
            # textBoxes = self.ocr.textboxSeekerTrainer(tem, 50)

            # print(text)
            
        elif len(self.recoredArea) > 0 and self.testNum == 4 and self.test:
            print(self.landMark)
            self.scollingRecognizeController(self.testScrollingArea, self.landMarkPosition, \
                (self.testScrollingArea, 3, 0, True), self.landMark, "downward", 1) # tem, 3, self.testModelID, True
            self.testNum = -1
    def eventAction(self) -> None:

        if self.statusID == 2000:
            if self.keyBoardInterrupt.statusGet() == 2:
                # print("ssssssssssssssss",self.statusID)
                self.x = -10
                self.y = -10
                self.screenShot = 2

                for x in self.__rec:
                    transformed = self.transform(x)
                    # if self.counter == 0:
                    #     self.textAreUpperBound = min(transformed["top"] , self.textAreUpperBound)
                    #     self.textAreLowerBound = max(transformed["top"] + transformed["height"],\
                    #         self.textAreLowerBound)
                    # print(self.textAreUpperBound, self.textAreLowerBound)
                    
                    self.recoredArea.append(transformed)
                    print(transformed)
                self.widgetsCleaner(self.__canvas, self.screenShotor)
                self.__rec.clear()
                self.width, self.height, self.positionX, self.positionY = self.windowSize["root"]
                # self.__root.attributes("-alpha", self.alpha)
                # if self.__subWindows != None:
                #     self.__subWindows.attributes("-alpha", self.alpha)
                # self.counter += 1
                
                self.statusID = -1
                self.messageBox = showinfo("RseMessager", "Please check if the result looks \
correct, please also clear the target area to ensure model could catch the right area.(Press key 'alt+z' finish the step)")
                tem = self.recoredArea.pop() # Show result
                # text=self.ocr.areTextTransfer(tem)
                # textBoxes = self.ocr.textboxSeekerTrainer(tem, 50)
                textBoxesParam = (tem, 3, self.testModelID, True)
                self.sequencialCommand.append(copy.deepcopy(textBoxesParam))
                
                self.testScrollingArea = copy.deepcopy(tem)
                self.textBoxes = self.ocr.textboxSeekerPredictor(tem, 3, self.testModelID, True)
                self.screenShotCreation()
                counter = 0
                for x in self.textBoxes:
                    leftX = (x["left"] * self.width) / self.screen.width  # 转换相对坐标。
                    leftY = (x["top"] * self.height) / self.screen.height

                    rightX = ((x["left"] + x["width"]) * self.width) / self.screen.width  # 转换相对坐标。
                    rightY = ((x["top"] + x["height"]) * self.height) / self.screen.height
                    if counter == 0:
                        self.rectangleCreation(leftX, leftY, rightX, rightY, "crimson")
                    elif counter == 1:
                        self.rectangleCreation(leftX, leftY, rightX, rightY, "blue")
                    else:
                        self.rectangleCreation(leftX, leftY, rightX, rightY, "green")
                    counter += 1
                    if counter % 3 == 0:
                        counter = 0
                self.statusID = 3000
                
        elif self.statusID == 2001:
            if self.keyBoardInterrupt.statusGet() == 2:
                # print("ssssssssssssssss",self.statusID)
                self.x = -10
                self.y = -10
                self.screenShot = 2

                for x in self.__rec:
                    transformed = self.transform(x)
                    # if self.counter == 0:
                    #     self.textAreUpperBound = min(transformed["top"] , self.textAreUpperBound)
                    #     self.textAreLowerBound = max(transformed["top"] + transformed["height"],\
                    #         self.textAreLowerBound)
                    # print(self.textAreUpperBound, self.textAreLowerBound)
                    
                    self.recoredArea.append(transformed)
                    print(transformed)
                self.widgetsCleaner(self.__canvas, self.screenShotor)
                self.__rec.clear()
                self.width, self.height, self.positionX, self.positionY = self.windowSize["root"]
                
                self.statusID = -1
                tem = copy.deepcopy(self.recoredArea)
                
                textBoxesParam = (tem, {"top": 0, "left": 0, "width": \
                self.screen.width, "height": self.screen.height}, self.testModelID)
                
                self.sequencialCommand.append(copy.deepcopy(textBoxesParam))
                
                testResult = self.ocr.textLocationPredictor(tem, {"top": 0, "left": 0, "width": \
                self.screen.width, "height": self.screen.height}, self.testModelID)
                
                if testResult:
                    self.messageBox = askquestion("RseMessager", "Model determine the edges of target \
area do not cut any text, it is a safe target area. Is it correct?")
                else:
                    self.messageBox = askquestion("RseMessager", "Model determine the edges of target \
area cut some text. Is it correct?")
                
                if self.messageBox == "yes":
                    pass
                
                self.recoredArea.clear()
                
                self.__root.attributes("-alpha", self.alpha)
                if self.__subWindows != None:
                    self.__subWindows.attributes("-alpha", self.alpha)
                self.statusID = 6202
                self.Start(6202)
            
        elif self.statusID == 2002: # find out text
            if self.keyBoardInterrupt.statusGet() == 2:
                # print("ssssssssssssssss",self.statusID)
                self.x = -10
                self.y = -10
                self.screenShot = 2
                for x in self.__rec:
                    transformed = self.transform(x)
                    if self.counter == 0:
                        self.textAreUpperBound = min(transformed["top"] , self.textAreUpperBound)
                        self.textAreLowerBound = max(transformed["top"] + transformed["height"],\
                            self.textAreLowerBound)
                        self.counter += 1
                    print(self.textAreUpperBound, self.textAreLowerBound)
                    
                    self.recoredArea.append(transformed)
                    
                self.currentColumnNum = len(self.recoredArea)
                
                self.widgetsCleaner(self.__canvas, self.screenShotor)
                self.__rec.clear()
                self.width, self.height, self.positionX, self.positionY = self.windowSize["root"]
                
                self.statusID = -1
                result = []
                textBoxesParam = (self.recoredArea, self.textBoxes, {"top": 0, "left": 0, "width": \
                    self.screen.width, "height": self.screen.height}, self.testModelID, "predict", \
                        self.textAreLowerBound - self.textAreUpperBound, self.textAreUpperBound, \
                                    self.textAreLowerBound)
                
                self.sequencialCommand.append(copy.deepcopy(textBoxesParam))
                
                for x in self.textBoxes:
                    tem = copy.deepcopy(self.recoredArea)
                    if len(self.ocr.relativeDistance[self.testModelID]) == 0:
                        print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
                        result += self.ocr.textSeeker(tem, copy.deepcopy(x), {"top": 0, "left": 0, "width": \
                            self.screen.width, "height": self.screen.height}, self.testModelID, "train", \
                                self.textAreLowerBound - self.textAreUpperBound, self.textAreUpperBound, \
                                    self.textAreLowerBound)
                    else:
                        print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
                        result += self.ocr.textSeeker(tem, copy.deepcopy(x), {"top": 0, "left": 0, "width": \
                            self.screen.width, "height": self.screen.height}, self.testModelID, "predict")
                
                ocrParam = (self.testScrollingArea, result, True, \
                    self.imageFilterOut, self.indexFilterOut, (0,0,self.screen.width,self.screen.height))
                
                self.sequencialCommand.append(copy.deepcopy(ocrParam))
                textMatched = self.ocr.areTextTransfer(self.testScrollingArea, result, True, \
                    self.imageFilterOut, self.indexFilterOut, (0,0,self.screen.width,self.screen.height))
                
                self.screenShotCreation()
                counter = 0
                formCounter = 0
                del self.temForm
                self.temForm = dict()
                for x in range(len(result)):
                    leftX = (result[x]["left"] * self.width) / self.screen.width  # 转换相对坐标。
                    leftY = (result[x]["top"] * self.height) / self.screen.height

                    rightX = ((result[x]["left"] + result[x]["width"]) * self.width) / self.screen.width  # 转换相对坐标。
                    rightY = ((result[x]["top"] + result[x]["height"]) * self.height) / self.screen.height
                    if counter == 0:
                        self.rectangleCreation(leftX, leftY, rightX, rightY, "crimson")
                    elif counter == 1:
                        self.rectangleCreation(leftX, leftY, rightX, rightY, "blue")
                    else:
                        self.rectangleCreation(leftX, leftY, rightX, rightY, "green")
                    if formCounter not in self.temForm.keys():
                        self.temForm[formCounter] = list()
                        
                    self.temForm[formCounter].append(textMatched[x])
                    counter += 1
                    formCounter += 1
                    if counter % 3 == 0:
                        counter = 0
                    if formCounter % self.currentColumnNum == 0:
                        formCounter = 0
                
                if len(self.temForm.keys()) > 0:
                    tem = []
                    for x in range(len(self.temForm.keys())):
                        tem.append(self.temForm[x][-2])
                self.landMark = tem
                self.landMarkPosition = self.textBoxes[-2]
                self.displacementMouse = 0
                print(self.sequencialCommand)
                self.statusID = 3001
        
        elif self.statusID == 2003: # extra step text matching
            if  self.keyBoardInterrupt.statusGet()== 2:
                self.x = -10
                self.y = -10
                self.screenShot = 2

                for x in self.__rec:
                    transformed = self.transform(x)
                    
                    self.recoredArea.append(transformed)
                    print(transformed)
                self.widgetsCleaner(self.__canvas, self.screenShotor)
                self.__rec.clear()
                self.width, self.height, self.positionX, self.positionY = self.windowSize["root"]
                
                self.messageBox = askquestion("RseMessager", "Is the game area same with the form?")
                if self.messageBox == "yes":
                    textMatched = self.ocr.areTextTransfer(self.testScrollingArea, copy.deepcopy(self.recoredArea), True)
                else:
                    textMatched = self.ocr.areTextTransfer({"top": 0, "left": 0, "width": \
                        self.screen.width, "height": self.screen.height}, copy.deepcopy(self.recoredArea), True)
                
                self.messageBox = askquestion("RseMessager", "Are the results correct? result: " + str(textMatched))
                
                if len(textMatched) == 1:
                    self.temForm[self.currentColumnNum] = copy.deepcopy(textMatched)
                    self.currentColumnNum += 1
                elif len(textMatched) > 1:
                    for x in range(len(textMatched)):
                        self.temForm[self.currentColumnNum] = []
                        self.temForm[self.currentColumnNum].append(textMatched[x])
                        self.currentColumnNum += 1
                
                self.recoredArea.clear()
                self.__root.attributes("-alpha", self.alpha)
                if self.__subWindows != None:
                    self.__subWindows.attributes("-alpha", self.alpha)
                    
                self.statusID = -1
                self.Start(6406)
        
        elif self.statusID == 2004:
            if  self.keyBoardInterrupt.statusGet()== 2:
                self.x = -10
                self.y = -10
                self.screenShot = 2

                for x in self.__rec:
                    transformed = self.transform(x)
                    
                    self.recoredArea.append(transformed)
                    print(transformed)
                self.widgetsCleaner(self.__canvas, self.screenShotor)
                self.__rec.clear()
                self.width, self.height, self.positionX, self.positionY = self.windowSize["root"]
                
                textMatched = self.ocr.areTextTransfer(copy.deepcopy(self.recoredArea[-1]))
                
                self.messageBox = askquestion("RseMessager", "Are the results correct? result: " + str(textMatched))
                
                if len(textMatched) == 1:
                    self.temForm[self.currentColumnNum] = copy.deepcopy(textMatched)
                    self.currentColumnNum += 1
                elif len(textMatched) > 1:
                    for x in range(len(textMatched)):
                        self.temForm[self.currentColumnNum] = []
                        self.temForm[self.currentColumnNum].append(textMatched[x])
                        self.currentColumnNum += 1
                
                self.recoredArea.clear()
                self.__root.attributes("-alpha", self.alpha)
                if self.__subWindows != None:
                    self.__subWindows.attributes("-alpha", self.alpha)
                    
                self.statusID = -1
                self.Start(6406)
        
        elif self.statusID == 2010: # record click
            # print(self.userInteraction.recoredBehaviours)
            clickedX, clickedY = self.listener.mouseGet("left")
            if clickedX != -1 and clickedY != -1:
                print(clickedX, clickedY)
                self.userInteraction.actionRecord(self.description[self.statusID],\
                    (clickedX, clickedY), self.recordPointer)
                self.shownListManagement("add", self.description[self.statusID]+": "+\
                        str(clickedX)+","+str(clickedY))
                
                print(self.userInteraction.recoredBehaviours)
                self.__root.attributes("-alpha", self.alpha)
                if self.__subWindows != None:
                    self.__subWindows.attributes("-alpha", self.alpha)
                self.statusID = 1020
                
        elif self.statusID == 3000: # stop showing result
            if self.keyBoardInterrupt.statusGet() == 2:
                # print("ssssssssssssssss",self.statusID)
                self.x = -10
                self.y = -10
                self.screenShot = 2

                self.widgetsCleaner(self.__canvas, self.screenShotor)
                self.__rec.clear()
                self.recoredArea.clear()
                self.__root.attributes("-alpha", self.alpha)
                if self.__subWindows != None:
                    self.__subWindows.attributes("-alpha", self.alpha)
                
                self.statusID = 6200
                self.Start(6200)
                
        elif self.statusID == 3001: # stop showing result
            if self.keyBoardInterrupt.statusGet()== 2:
                # print("ssssssssssssssss",self.statusID)
                self.x = -10
                self.y = -10
                self.screenShot = 2

                self.widgetsCleaner(self.__canvas, self.screenShotor)
                self.__rec.clear()
                self.recoredArea.clear()
                # self.__root.attributes("-alpha", self.alpha)
                # if self.__subWindows != None:
                #     self.__subWindows.attributes("-alpha", self.alpha)
                # self.__root.geometry("{0}x{1}+{2}+{3}" \
                #     .format(self.windowSize["record"][0], self.windowSize["record"][1], \
                #         self.windowSize["record"][2], self.windowSize["record"][3]))
                # self.__subWindows.geometry("{0}+{1}" \
                #     .format(self.subWindowInfo[2], self.subWindowInfo[3]))
                
                self.messageBox = askquestion("RseMessager", "Are the results correct?")
                
                if self.messageBox == "yes":
                    self.dbManagement.OCRModelDataSaver(self.ocr.centroids,\
                        self.ocr.data, self.ocr.modelID, self.ocr.relativeDistance)
                
                
                self.imageFilterOut, self.indexFilterOut = None, None
                self.statusID = 7000
                self.Start(7100)
                
        elif 5000 > self.statusID >= 4201: # execute recorded command
            self.__root.attributes("-alpha", 0)
            if self.__subWindows != None:
                self.__subWindows.attributes("-alpha", 0)
                
            print(self.userInteraction.recoredBehaviours)
            print(len(self.userInteraction.recoredBehaviours[self.statusID % 4201]))
            print(self.executePointer)
            
            if len(self.userInteraction.recoredBehaviours) == 0 or self.executePointer >= \
                len(self.userInteraction.recoredBehaviours[self.statusID % 4201]):
                    
                self.__root.attributes("-alpha", self.alpha)
                if self.__subWindows != None:
                    self.__subWindows.attributes("-alpha", self.alpha)
                self.executePointer = 0
                self.loopCounter = ([],[],[],[])
                self.statusID = 4000
            else:
                command, param = self.userInteraction.actionExcute(self.statusID % 4201, self.executePointer)
                command(*param)
                self.executePointer += 1
                
            if self.keyBoardInterrupt.statusGet() == 2:
                self.__root.attributes("-alpha", self.alpha)
                if self.__subWindows != None:
                    self.__subWindows.attributes("-alpha", self.alpha)
                self.executePointer = 0
                self.loopCounter = ([],[],[],[])
                self.statusID = 4000
        
        elif 7004 > self.statusID >= 7001:
            self.counter += 1
            if self.keyBoardInterrupt.statusGet()== 2:
                self.IOController.mouse.position = (self.screen.width - 1, self.screen.height + 1)
                self.counter = 0
                self.textBoxes = self.ocr.textboxSeekerPredictor(*self.sequencialCommand[0])
                screen = np.array(self.ocr.sct.grab(self.testScrollingArea))
                image_rgb = cv2.cvtColor(screen, cv2.COLOR_BGR2RGB)

                result = self.ocr.ocr.ocr(image_rgb, cls=False)
                # print(result)
                txts = []
                boxes = []
                for x in result:
                    boxes = [line[0] for line in x]
                    txts = [line[1][0] for line in x]
                    
                landMarkUnCheckIndex = list(np.arange(len(self.landMark)))
                targetBoxes = []
                for x in range(len(txts)):
                    counter = 0
                    while counter < len(landMarkUnCheckIndex):
                        print(Levenshtein.distance(txts[x], self.landMark[counter], score_cutoff=1))
                        if Levenshtein.distance(txts[x], self.landMark[counter], score_cutoff=1) <= 1:
                            print(self.landMark[counter])
                            landMarkUnCheckIndex.pop(counter)
                            targetBoxes.append(boxes[x])
                            counter -= 1
                        counter += 1
                    if counter == 0:
                        break
                textBoxesResult = None
                for x in range(len(self.textBoxes)):
                    counter = 0
                    for y in range(len(targetBoxes)):
                        overlap_area = self.ocr.calculate_overlap([self.textBoxes[x]["left"], \
                            self.textBoxes[x]["top"], self.textBoxes[x]["left"] + self.textBoxes[x]["width"], \
                                self.textBoxes[x]["top"] + self.textBoxes[x]["height"]], \
                                    targetBoxes[y][0] + targetBoxes[y][2])
                        print(overlap_area)
                        print([self.textBoxes[x]["left"], \
                            self.textBoxes[x]["top"], self.textBoxes[x]["left"] + self.textBoxes[x]["width"], \
                                self.textBoxes[x]["top"] + self.textBoxes[x]["height"]], \
                                    targetBoxes[y][0] + targetBoxes[y][2])
                        print("++++++++++++++++++++++++")
                        if overlap_area > 0.9 * (targetBoxes[y][2][0] - targetBoxes[y][0][0]) * \
                            (targetBoxes[y][2][1] - targetBoxes[y][0][1]):
                            counter += 1
                    if counter >= len(targetBoxes):
                        textBoxesResult = self.textBoxes[x]
                        break
                        
                if textBoxesResult != None:
                    print(textBoxesResult["top"] - self.landMarkPosition["top"])
                    print(self.displacementMouse)
                    print(self.displacementMouse/(textBoxesResult["top"] - self.landMarkPosition["top"]))
                    print((self.displacementMouse - (textBoxesResult["top"] - self.landMarkPosition["top"])) * self.displacementCount)
                self.statusID += 1
                    
                # self.sequencialCommand[2][0], textbox, self.sequencialCommand[2][2], self.sequencialCommand[2][3], \
                #     "predict", self.sequencialCommand[2][5], self.sequencialCommand[2][6], self.sequencialCommand[2][7]
                if 7004 > self.statusID >= 7001:
                    result = []
                    for x in self.textBoxes[1:-1]:
                        if len(self.ocr.relativeDistance[self.testModelID]) == 0:
                            print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
                            result += self.ocr.textSeeker(self.sequencialCommand[2][0], copy.   deepcopy(x),\
                                self.sequencialCommand[2][2], self.sequencialCommand[2][3],     "train", \
                                    self.sequencialCommand[2][5], self.sequencialCommand[2][6], \
                                        self.sequencialCommand[2][7])
                        else:
                            print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
                            result += self.ocr.textSeeker(self.sequencialCommand[2][0], copy.   deepcopy(x), \
                                self.sequencialCommand[2][2], self.sequencialCommand[2][3],     "predict")


                    textMatched = self.ocr.areTextTransfer(self.testScrollingArea, result, True, \
                        self.imageFilterOut, self.indexFilterOut, (0,0,self.screen.width,self.  screen.height))


                    tem = []
                    for x in range(len(self.landMark)):
                        tem.append(textMatched[-1 * len(self.landMark) + x])
                    self.landMark = tem
                    self.landMarkPosition = self.textBoxes[-1]
                    self.displacementMouse = 0
                    self.displacementCount = 0

                    tem = self.statusID
                    self.statusID = -1
                    self.messageBox = showinfo("RseMessager", "This is our next test: \n \
    Data:\n \
        " + str(self.landMark))
                    self.statusID = tem
                
            elif self.counter % 3 == 0:
                self.counter = 0
                self.IOController.mouse.position = ((self.testScrollingArea["left"] + self.testScrollingArea["width"])/2 ,\
                    self.testScrollingArea["top"] + int(self.testScrollingArea["height"] * 9/10))
                time.sleep(0.1)
                self.IOController.mouse.press(Button.left)
                self.IOController.smooth((self.testScrollingArea["left"] + self.testScrollingArea["width"])/2 ,\
                    self.testScrollingArea["top"] + int(self.testScrollingArea["height"] * 8/10))
                time.sleep(0.1)
                self.IOController.mouse.release(Button.left)
                self.displacementMouse += int(self.testScrollingArea["height"] * 1/10)
                self.displacementCount += 1
                time.sleep(0.1)
            
            
        elif self.statusID == 7004:
            self.__root.attributes("-alpha", self.alpha)
            if self.__subWindows != None:
                self.__subWindows.attributes("-alpha", self.alpha)
            self.__root.geometry("{0}x{1}+{2}+{3}" \
                .format(self.windowSize["record"][0], self.windowSize["record"][1], \
                    self.windowSize["record"][2], self.windowSize["record"][3]))
            self.__subWindows.geometry("{0}x{1}+{2}+{3}" \
                .format(self.subWindowInfo[0], self.subWindowInfo[1], self.subWindowInfo[2], self.subWindowInfo[3]))
            
            self.statusID = -1
            self.Start(6402)
    
    def widgetsCleaner(self, *arrayLike:Iterable) -> None:
        for x in arrayLike:
            if isinstance(x, Iterable):
                for y in x:
                    if y != None:
                        y.destroy()
                x.clear()
            else:
                if x != None:
                    x.destroy()
    
    def shownListManagement(self, mode:str, newData:str, indexTo = -1):
        if mode == "add":
            if indexTo == -1:
                self.mylistBox.insert(tkinter.END, str(self.mylistBox.size() + 1)+". "+newData)
            else:
                self.mylistBox.insert(indexTo, newData)
        elif mode == "del":
            if indexTo == -1:
                self.mylistBox.delete(tkinter.END)
            else:
                self.mylistBox.delete(indexTo)
        
    def waitUntilIconDetected(self, image, region: tuple[int]):
        try:
            if not self.IOController.iconDetection(image, region):
                self.executePointer -= 1
        except Exception as e:
            print("error occured, wait for a moment and try again.")
            print(e)
            self.executePointer -= 1
            
    def timeWait(self, timeCount:int):
        if self.counter == -1:
            self.counter = time.time()
        if time.time() - self.counter < timeCount:
            self.executePointer -= 1
        else:
            self.counter = -1

    def loopController(self, loopBackTo:int, loopTime:int, loopID:int):
        print(self.loopCounter)
        if len(self.loopCounter[loopID]) == 0:
            self.loopCounter[loopID].append(loopTime)
        if self.loopCounter[loopID][0] > 0 or self.loopCounter[loopID][0] == -1:
            self.executePointer = copy.deepcopy(loopBackTo) - 1
            self.loopCounter[loopID][0] -= int(loopTime > 0)

    def drawer(self) -> int:
        if self.listener != None:


            if self.screenShot == 1 and self.__num > 0:  # 1是截图功能的id，number是多少个画了多少个矩形。
                temx, temy = self.listener.mouseGet("left")  # 鼠标的绝对坐标。

                temx = (temx * self.width) / self.screen.width  # 转换相对坐标。
                temy = (temy * self.height) / self.screen.height

                # print("call:",temx,temy)
                if self.x == -10 or self.y == -10:  # prevent first click detection
                    self.x = -1
                    self.y = -1
                elif (temx != self.x or temy != self.y) and (temx != -1 and temy != self.y):  # 防止长度和宽度为0的矩形。
                    print("click:", temx, temy)
                    self.x, self.y = temx, temy
                    self.__counter += 1  # 初始是0，每点击一次加1。
                    if self.__counter == 1:  # 创捷一个新的正方形。
                        self.rectangleCreation(self.x, self.y, self.x, self.y, width=3)
                    else:
                        self.__counter = 0

                if self.__counter == 1:  # 正方形点击第一下该如何反应。
                    temx, temy = self.listener.motionGet()  # 检测鼠标移动坐标（绝对坐标）

                    temx = (temx * self.width) / self.screen.width  # 转换成相对坐标
                    temy = (temy * self.height) / self.screen.height

                    if (self.xRight == -1 and self.yRight == -1) \
                            or (temx != self.xRight or temy != self.yRight):  # xRight和yRight检测前一刻和后一刻一样不一样

                        print("move:", temx, temy)
                        self.xRight, self.yRight = temx, temy  # 鼠标当前位置
                        self.rectangleConfigure(self.x, self.y, self.xRight, self.yRight, width=3)  # 更新矩阵

        if self.x == 0 or self.y == 0:  # 紧急退出
            self.listener.terminate()
            self.__root.destroy()
            return -1

        return 1

    def canvasPlace(self, positionX=0, positionY=0, highlightthickness=0, bgColor="black", target="root") -> None:

        if target == "sub":
            self.__canvas = tkinter.Canvas(self.screenShotor, highlightthickness= \
                highlightthickness, width=self.width, height=self.height, bg=bgColor)
        else:
            self.__canvas = tkinter.Canvas(self.__root, highlightthickness= \
                highlightthickness, width=self.width, height=self.height, bg=bgColor)
        self.__canvas.place(x=positionX, y=positionY)

    def rectangleCreation(self, positionX=0, positionY=0, rightX=0, rightY=0, outline="crimson", \
                          width=3, dash=(1, 1)) -> None:

        self.__rec.append(self.__canvas.create_rectangle(positionX, positionY, rightX, rightY, \
                                                         outline=outline, width=width, dash=dash))
        print("coor:", self.__canvas.coords(self.__rec[-1]))

    def rectangleConfigure(self, positionX=0, positionY=0, rightX=0, rightY=0, outline="crimson", \
                           width=3, index=-1, dash=(1, 1)) -> None:

        self.__canvas.itemconfigure(self.__rec[index], outline=outline, width=width, \
                                    dash=dash)

        self.__canvas.coords(self.__rec[index], positionX, positionY, rightX, rightY)
        print("coorMoving:", self.__canvas.coords(self.__rec[-1]))

        # self.__canvas.itemconfigure(self.__rec[index], outline=outline, width=width, \
        #                             dash=dash)

        # self.__canvas.coords(self.__rec[index], positionX, positionY, rightX, rightY)
        # print("coorMoving:", self.__canvas.coords(self.__rec[-1]))


    def transform(self, canvas_rectangle):  # 将画布上的相对坐标转换成屏幕的绝对坐标
        coords = self.__canvas.coords(canvas_rectangle)  # 得到矩阵的坐标
        tem_left_x = coords[0]
        tem_left_y = coords[1]
        tem_right_x = coords[2]
        tem_right_y = coords[3]

        tem_left_x = tem_left_x * self.screen.width / self.width  # 转换成绝对坐标
        tem_left_y = tem_left_y * self.screen.width / self.width
        tem_right_x = tem_right_x * self.screen.width / self.width
        tem_right_y = tem_right_y * self.screen.width / self.width

        monitor = {
            "top": int(tem_left_y),
            "left": int(tem_left_x),
            "width": int(tem_right_x - tem_left_x),
            "height": int(tem_right_y - tem_left_y)
        }

        return monitor

    def listRecognizeController(self, textBoxesParm):
        self.textBoxes = self.ocr.textboxSeekerPredictor(*textBoxesParm) # tem, 3, self.testModelID, True
        

    def scollingRecognizeController(self, scrollingArea:dict, landMarkCoord:dict, textBoxesParm: tuple, \
        landMark:list[str] = None, direction = "upward", scrollingRatio = 1, scrollingEqualDelta = 0):
        # textBoxesParam = (self.recoredArea, self.textBoxes, {"top": 0, "left": 0, "width": \
        #     self.screen.width, "height": self.screen.height}, self.testModelID, "predict")
        # landMark = None
        # textNum = len(textSeekerParm[1])
        # self.textBoxes = self.ocr.textboxSeekerPredictor(*textBoxesParm) # tem, 3, self.testModelID, True
        # if len(self.textBoxes) >= 2:
        #     if self.textBoxes[0]["top"] > textBoxesParm[0]["top"] + self.textBoxes[0]["height"]:
        #         result = []
        #         for x in self.textBoxes[:-1]:
        #             # text positions, textBoxes, screenArea, testModelID, mode
        #             result += self.ocr.textSeeker(*textSeekerParm)
                
                
        #         textMatched = self.ocr.areTextTransfer(textBoxesParm[0], result, True, \
        #             imageFilterOut, imageFilterOut, (0,0,self.screen.width,self.screen.height))
        #         landMark = textMatched[:-1*textNum]
        print("Scrolling...")
        if landMark == None:
            self.textBoxes = self.ocr.textboxSeekerPredictor(*textBoxesParm) # tem, 3, self.testModelID, True
            if len(self.textBoxes) >= 2:
                return True
            else:
                while len(self.textBoxes) < 2:
                    if len(self.textBoxes) > 0:
                        scrollDistance = int(self.textBoxes[0]["height"])
                    else:
                        scrollDistance = int(scrollingArea["height"]/10)
                    self.IOController.mouse.position = ((scrollingArea["left"] + scrollingArea["width"])/2 ,\
                        scrollingArea["top"] + scrollingArea["height"] - int(scrollingArea["height"]/10))
                    time.sleep(0.1)
                    self.IOController.mouse.press(Button.left)
                    self.IOController.smooth((scrollingArea["left"] + scrollingArea["width"])/2 ,\
                        scrollingArea["top"] + scrollingArea["height"] - int(scrollingArea["height"]/10) \
                            - scrollDistance)
                    time.sleep(0.1)
                    self.IOController.mouse.release(Button.left)
                    time.sleep(0.1)
                return True
        else:
            landMarkUnCheckIndex = list(np.arange(len(landMark)))
            scrollCounut = 1
            positonBias = int(scrollingArea["height"]/10)
            scrollDistance = int((landMarkCoord["top"] - scrollingArea["top"])/scrollingRatio \
                + scrollingEqualDelta)
            while scrollingArea["height"]*4/5 < scrollDistance:
                scrollDistance = scrollDistance / 2
                scrollCounut += 1
            scrollDistance = int(scrollDistance)
            for x in range(scrollCounut):
                self.IOController.mouse.position = ((scrollingArea["left"] + scrollingArea["width"])/2 ,\
                    scrollingArea["top"] + scrollingArea["height"] - int(scrollingArea["height"]/10))
                time.sleep(0.1)
                self.IOController.mouse.press(Button.left)
                self.IOController.smooth((scrollingArea["left"] + scrollingArea["width"])/2 ,\
                    scrollingArea["top"] + scrollingArea["height"] - int(scrollingArea["height"]/10) \
                        - scrollDistance)
                time.sleep(0.1)
                self.IOController.mouse.release(Button.left)
                time.sleep(0.1)
            
            while len(landMarkUnCheckIndex) > 0:
                self.keyBoardInterrupt.activeFlagSet(1)
                self.listener.activeFlagSet(1)
                if self.keyBoardInterrupt.statusGet() == 2:
                    return False
                
                print(landMarkUnCheckIndex)
                allText = self.ocr.areTextTransfer(scrollingArea)
                for x in range(len(allText)):
                    counter = 0
                    while counter < len(landMarkUnCheckIndex):
                        print(Levenshtein.distance(allText[x], landMark[counter], score_cutoff=1))
                        if Levenshtein.distance(allText[x], landMark[counter], score_cutoff=1) <= 1:
                            print(landMark[counter])
                            landMarkUnCheckIndex.pop(counter)
                            counter -= 1
                        counter += 1
                    if counter == 0:
                        return True
                    
                self.IOController.mouse.position = ((scrollingArea["left"] + scrollingArea["width"])/2 ,\
                    scrollingArea["top"] + int(scrollingArea["height"]/10))
                time.sleep(0.1)
                self.IOController.mouse.press(Button.left)
                self.IOController.smooth((scrollingArea["left"] + scrollingArea["width"])/2 ,\
                    scrollingArea["top"] + int(scrollingArea["height"]/5)/scrollingRatio \
                        + scrollingEqualDelta)
                time.sleep(0.1)
                self.IOController.mouse.release(Button.left)
                time.sleep(0.1)

            

class mouse_control():
    def __init__(self):
        self.mouse = Controller()

    def move_and_press_mouse(self, x_position, y_position):
        self.mouse.position = (x_position, y_position)  # Move the mouse to the position
        time.sleep(0.1)
        self.mouse.press(Button.left)
        time.sleep(0.1)
        self.mouse.release(Button.left)

    def smooth(self, x_position, y_position, speed=150):
        position_xy = pyautogui.position()
        distance = math.sqrt(pow(x_position - position_xy[0], 2) + pow(y_position - position_xy[1], 2))
        print(distance)
        mouse_time = distance / speed
        print(mouse_time)
        pyautogui.moveTo(x_position, y_position, mouse_time)
        
    def clickOnButton(self, image, region: tuple[int]) -> None:
        print(type(image))
        # image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        print(region)
        help_pos = pyautogui.locateOnScreen(image,confidence=0.7,region=region, minSearchTime = 1) #region中的参数为xy起始点，宽度和高度
        print("----------------")
        print(help_pos)
        if help_pos != None:
            # print(help_pos)
            goto_pos = pyautogui.center(help_pos) # 找到传回图片的中心点,并传回坐标
            self.move_and_press_mouse(goto_pos.x, goto_pos.y)
            print(goto_pos)
            
    def iconDetection(self, image, region: tuple[int]) -> bool:
        # image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

        help_pos = pyautogui.locateOnScreen(image,confidence=0.7,region=region, minSearchTime = 1) #region中的参数为xy起始点，宽度和高度
        # print("----------------")
        # print(help_pos)
        if help_pos != None:
            return True
        else:
            return False


class OCRController():
    def __init__(self, path, incodeFile) -> None:
        # pytesseract.pytesseract.tesseract_cmd = r""
        self.sct = mss()
        self.currentPath = path
        self.ocr = PaddleOCR(use_angle_cls = True, use_gpu = False, lang = "ch",
                rec_model_dir = self.currentPath + "\\inference\\ch_PP-OCRv4_rec_infer\\",
                cls_model_dir = self.currentPath + "\\inference\\ch_ppocr_mobile_v2.0_cls_infer\\",
                det_model_dir = self.currentPath + "\\inference\\ch_PP-OCRv4_det_infer\\") 

        # print(incodeFile)
        self.kmeans = None
        self.centroids, self.data ,self.modelID, self.relativeDistance = incodeFile
        self.centroids, self.data ,self.modelID = np.array(self.centroids), \
            np.array(self.data), np.array(self.modelID)
        self.temResult = []
        self.modelLabels = []
        self.relativeDistance = list(self.relativeDistance)
        
        self.screen = np.array([])
        
        arrivedResult = []
        # print(self.modelLabels, self.centroids, self.data ,self.modelID, self.relativeDistance)
        
    def __closest(self, a):
        return np.argmin(a)
    
    def calculate_overlap(self, box1, box2):
        # Calculate the overlapping area between two text boxes
        x1 = max(box1[0], box2[0])
        y1 = max(box1[1], box2[1])
        x2 = min(box1[2], box2[2])
        y2 = min(box1[3], box2[3])

        width = max(0, x2 - x1)
        height = max(0, y2 - y1)

        overlap_area = width * height
        return overlap_area
    
    def areTextTransfer(self, targetArea:dict, textPosition:list[dict] = [], match = False, \
        imageFilterOut:np.ndarray = None, position:int = None, IconArea:tuple = None) -> list[str]:
        
        if np.any(imageFilterOut != None):
            coos = pyautogui.locateAllOnScreen(imageFilterOut, confidence=0.7,region=IconArea)
            coos = list(coos)
        screen = np.array(self.sct.grab(targetArea))
        image_rgb = cv2.cvtColor(screen, cv2.COLOR_BGR2RGB)

        result = self.ocr.ocr(image_rgb, cls=False)
        # print(result)
        txts = []
        for x in result:
            if match:
                boxes = [line[0] for line in x]
            txts = [line[1][0] for line in x]
            
        if match:
            print(txts)
            print(boxes)
            print(textPosition)
            print(">>>>>>>>>>>>>>>>>>>>")
            matched_boxes = []
            filterIndex = []

            for list_box in range(len(textPosition)):
                max_overlap = 0
                matched_box = None

                for screen_box in range(len(boxes)):
                    tem = textPosition[list_box].copy()
                    overlap_area = self.calculate_overlap([tem["left"] - targetArea["left"], \
                        tem["top"] - targetArea["top"], tem["left"] - targetArea["left"] + tem["width"], \
                            tem["top"] - targetArea["top"] + tem["height"]], \
                            boxes[screen_box][0] + boxes[screen_box][2])
                    if overlap_area > max_overlap:
                        max_overlap = overlap_area
                        matched_box = screen_box
                        
                    if list_box == 0:
                        if position != None and np.any(imageFilterOut != None):
                            lastx = None
                            lasty = None
                            for i in coos:
                                goto_pos = pyautogui.center(i)#找到传回图片的中心点,并传回坐标
                                if lastx == None or lasty == None or abs(lastx - goto_pos.x) \
                                    >= i.width or abs(lasty - goto_pos.y) >= i.height:
                                    #
                                    lastx, lasty = goto_pos.x, goto_pos.y
                                    wrongArea = self.calculate_overlap((i.left - targetArea["left"], i.top - targetArea["top"], i.left \
                                        + i.width - targetArea["left"], i.top + i.height - targetArea["top"]), \
                                            boxes[screen_box][0] + boxes[screen_box][2])
                                    if wrongArea >= 0.7 * i.width * i.height:
                                        if position != -1:
                                            txts[screen_box] = txts[screen_box][0:position] + txts[screen_box][position + 1:]
                                        else:
                                            print(boxes[screen_box])
                                            print(txts[screen_box])
                                            print(wrongArea)
                                            print(i)
                                            print("===========================")
                                            txts[screen_box] = txts[screen_box][0:position]

                if matched_box is not None:
                    matched_boxes.append(matched_box)
                else:
                    filterIndex.append(len(matched_boxes))
                    matched_boxes.append(0)
                    
            matched_boxes = np.array(matched_boxes, dtype = np.int32)
            targetText = np.array(txts)
            filterText = np.array(filterIndex, dtype = np.int32)
            print(matched_boxes, targetText)
            
            unProcessResult = targetText[matched_boxes]
            unProcessResult[filterText] = None
            print(unProcessResult)
            
            return list(unProcessResult)
        
        return txts
    
    def textboxSeekerTrainer(self, targetArea:dict, heightOfTarget:int) -> list[dict]:
        screen = np.array(self.sct.grab(targetArea))
        # print(screen.shape)
        # print(screen)
        gray_image = cv2.cvtColor(screen, cv2.COLOR_BGR2GRAY)
        # print(gray_image.shape)
        gray_image = np.array(gray_image, dtype = np.int32)
        
        self.KMEANSTrainer(gray_image, 2, 0, 500)
            
        class2 = np.where(self.modelLabels[-1] == 1)[0]
        result = []
        node = 0
        lastNode = 0
        for x in range(class2.shape[0]):
            if class2[x] > node:
                if node > (lastNode + 1):
                    tem=copy.deepcopy(targetArea)
                    tem["top"] = lastNode + targetArea["top"]
                    tem["height"] = node - lastNode
                    if heightOfTarget <= tem["height"]:
                        result.append(tem)
                    self.temResult.append(tem)
                    
                tem=copy.deepcopy(targetArea)
                tem["top"] = node + targetArea["top"]
                tem["height"] = class2[x] - node
                if heightOfTarget <= tem["height"]:
                    result.append(tem)
                self.temResult.append(tem)
                
                lastNode = copy.deepcopy(class2[x])
                node = copy.deepcopy(class2[x])
            node+=1
        self.temResult = copy.deepcopy(result)

        return result

        
    def __modeGetter(self,npArray):
        counts = np.bincount(npArray)
        
        return np.argmax(counts)

    def textboxSeekerPredictor(self, targetArea:dict, heightOfTarget:int, modelID:int, \
        meanThreshold = False) -> list[dict]:
        screen = np.array(self.sct.grab(targetArea))
        # print(screen.shape)
        # print(screen)
        gray_image = cv2.cvtColor(screen, cv2.COLOR_BGR2GRAY)
        # print(gray_image.shape)
        gray_image = np.array(gray_image, dtype = np.int32)
        
        labels = self.KMEANSPredictor(gray_image, modelID, 2, 500, 1)
        print(labels)
            
        class2 = np.where(labels == 1)[0]
        result = []
        self.temResult.clear()
        node = 0
        lastNode = 0
        heights = np.array([])
        for x in range(class2.shape[0]):
            if class2[x] > node:
                if node > (lastNode + 1):
                    tem=copy.deepcopy(targetArea)
                    tem["top"] = lastNode + targetArea["top"]
                    tem["height"] = node - lastNode

                    if heightOfTarget <= tem["height"]:
                        result.append(tem)
                        heights = np.append(heights, tem["height"])
                    self.temResult.append(tem)
                    
                tem=copy.deepcopy(targetArea)
                tem["top"] = node + targetArea["top"]
                tem["height"] = class2[x] - node
                if heightOfTarget <= tem["height"]:
                    result.append(tem)
                    heights = np.append(heights, tem["height"])
                self.temResult.append(tem)
                

                lastNode = copy.deepcopy(class2[x])
                node = copy.deepcopy(class2[x])
            node+=1
        
        if meanThreshold:
            result = np.array(result)
            threshold = np.mean(heights)
            strongTextBoxes = np.where(heights > threshold)[0]
            heights = heights[strongTextBoxes]
            result = result[strongTextBoxes]
            inlier = np.where(heights < threshold * 2.5)[0]
            result = result[inlier]
            result = list(result)
            
        if len(result) == 0:
            result = [copy.deepcopy(targetArea)]
        # self.temResult = copy.deepcopy(result)


        return result
    
    def unexpectedResult(self, targetArea:dict, heightOfTarget:int):
        lastTextbox = None
        for x in range(len(self.temResult)):
            pass
            
    def textLocationTrainer(self, textArea:list[dict], screenArea:dict):
        screen = np.array(self.sct.grab(screenArea))
        # print(screen.shape)
        # print(screen)
        gray_image = cv2.cvtColor(screen, cv2.COLOR_BGR2GRAY)
        # print(gray_image.shape)
        gray_image = np.array(gray_image, dtype = np.int32)
        
        boundaryPixel = np.array([], dtype = np.int32)
        for x in range(len(textArea)):
            boundaryPixel = np.append(boundaryPixel, gray_image[textArea[x]["top"], textArea[x]["left"] : \
                textArea[x]["left"] + textArea[x]["width"]])
            
            boundaryPixel = np.append(boundaryPixel, gray_image[textArea[x]["top"] + textArea[x]["height"], \
                textArea[x]["left"] : textArea[x]["left"] + textArea[x]["width"]])
        
        self.KMEANSTrainer(boundaryPixel, 3, 1, 600, False)
        
    
    def textLocationPredictor(self, textArea:list[dict], screenArea:dict, modelID:int) -> bool:
        screen = np.array(self.sct.grab(screenArea))
        # print(screen.shape)
        # print(screen)
        gray_image = cv2.cvtColor(screen, cv2.COLOR_BGR2GRAY)
        # print(gray_image.shape)
        gray_image = np.array(gray_image, dtype = np.int32)
        
        
        boundaryPixel = np.array([], dtype = np.int32)
        for x in range(len(textArea)):
            boundaryPixel = gray_image[textArea[x]["top"], textArea[x]["left"] : \
                textArea[x]["left"] + textArea[x]["width"]].copy()
            
            labelsUpper = self.KMEANSPredictor(boundaryPixel, modelID, 3, 600, 1,\
                False, x == 0)
            print(labelsUpper)
            labelsUpper = np.diff(labelsUpper)
            if np.where(labelsUpper == self.__get_mode(labelsUpper))[0].shape[0]\
                < labelsUpper.shape[0] - 2:
                return False
            
            
            boundaryPixel = gray_image[textArea[x]["top"] + textArea[x]["height"], \
                textArea[x]["left"] : textArea[x]["left"] + textArea[x]["width"]].copy()
            
            labelsLower = self.KMEANSPredictor(boundaryPixel, modelID, 3, 600, 1,\
                False, False)
            print(labelsLower)
            labelsLower = np.diff(labelsLower)
            if np.where(labelsLower == self.__get_mode(labelsLower))[0].shape[0] < labelsLower.shape[0] - 2:
                return False

        return True
        
        
    def __get_mode(self, Kdarray:np.ndarray):
        # Flatten the array into a 1-dimensional array
        flattened_array = np.ravel(Kdarray)

        # Get the unique values and their counts
        unique_values, counts = np.unique(flattened_array, return_counts=True)

        # Find the index of the maximum count
        max_count_index = np.argmax(counts)

        # Get the mode value(s) from the unique values array
        mode_value = unique_values[max_count_index]

        return mode_value

        
    def textSeeker(self, textArea:list[dict], textboxArea:dict, screenArea:dict, modelID:int, \
        mode = "train", textHeight = -1, upper = -1, lower = -1):
        if mode == "train":
            print(textboxArea, textHeight)
            print(screenArea)
            result = []
            for y in range(0, textboxArea["height"] - textHeight, 2):
                print(y)
                counter = 0
                result = []
                temRelativeDis = []
                for x in range(len(textArea)):
                    if y == 0:
                        textArea[x]["top"] = textArea[x]["top"] - (upper - textboxArea["top"])
                    else:
                        textArea[x]["top"] += 2
                    print(textArea[x])
                        
                    temRelativeDis.append(textArea[x]["top"] - textboxArea["top"])
                        
                    tem = self.areTextTransfer(textArea[x])
                    print(tem)
                    print("---------------------------")
                    if len(tem) > 0:
                        result.append(tem[0])
                    else:
                        counter += 1
                        result.append(None)
                        
                if counter <= len(result)/2 and self.textLocationPredictor(textArea, screenArea, modelID):
                    self.relativeDistance[modelID].append(temRelativeDis)
                    break
            print(result, "<<<<<<<<<<<<<<<<<<<<<<")
            # return result
            return textArea
        else:
            satisfyFlag = []
            trainList = []
            result = []
            resultUpper = []
            resultLower = []
            if textHeight == -1 or upper == -1 or lower == -1:
                upper = 90000
                lower = 0
            for y in range(len(self.relativeDistance[modelID])):
                result = []
                resultUpper = []
                resultLower = []
                counter = 0
                counterUpper = 0
                counterLower = 0
                adjustAreaLower = []
                adjustAreaUpper = []
                for x in range(len(textArea)):
                    if y == 0:
                        trainList.append(textArea[x].copy())
                        if textHeight == -1:
                            upper = min(textArea[x]["top"], upper)
                            lower = max(textArea[x]["top"] + textArea[x]["height"], lower)
                    textArea[x]["top"] = self.relativeDistance[modelID][y][x] + textboxArea["top"]
                    adjustAreaLower.append(textArea[x].copy())
                    adjustAreaLower[-1]["top"] += 2
                    adjustAreaUpper.append(textArea[x].copy())
                    adjustAreaUpper[-1]["top"] -= 2
                    

                    tem = self.areTextTransfer(adjustAreaUpper[-1])
                    print(tem)
                    if len(tem) > 0:
                        resultUpper.append(tem[0])
                    else:
                        counterUpper += 1
                        resultUpper.append(None)
                        
                    tem = self.areTextTransfer(textArea[x])
                    print(tem)
                    if len(tem) > 0:
                        result.append(tem[0])
                    else:
                        counter += 1
                        result.append(None)
                        
                        
                    tem = self.areTextTransfer(adjustAreaLower[x])
                    print(tem)
                    print("---------------------------")
                    if len(tem) > 0:
                        resultLower.append(tem[0])
                    else:
                        counterLower += 1
                        resultLower.append(None)
                        
                    
                if textHeight == -1:
                    textHeight = lower - upper
                    
                if self.textLocationPredictor(textArea, screenArea, modelID) and counter <= len(result)/2:
                    satisfyFlag.append(True)
                    break
                elif self.textLocationPredictor(adjustAreaUpper, screenArea, modelID) and counterUpper <= len(resultUpper)/2:
                    satisfyFlag.append(True)
                    break
                elif self.textLocationPredictor(adjustAreaLower, screenArea, modelID) and counterLower <= len(resultLower)/2:
                    satisfyFlag.append(True)
                    break
                else:
                    satisfyFlag.append(False)
                    
            print(satisfyFlag)
            if True in satisfyFlag:
                valid = satisfyFlag.index(True)
                if valid == 0:
                    # return result
                    print(result, "<<<<<<<<<<<<<<<<<<<<<<")
                    return textArea
                elif valid == 1:
                    # return resultUpper
                    print(resultUpper, "<<<<<<<<<<<<<<<<<<<<<<")
                    return adjustAreaUpper
                else:
                    # return resultLower
                    print(resultLower, "<<<<<<<<<<<<<<<<<<<<<<")
                    return adjustAreaLower
            else:
                return self.textSeeker(trainList, textboxArea, screenArea, modelID, "train", textHeight, upper, lower)

        
    def KMEANSTrainer(self, gray_image:np.ndarray, n_clusters = 2,random_state = 0, max_iter = 500, mode = True):
        
        if mode:
            patterns = np.apply_along_axis(self.__modeGetter, 1, gray_image) # modes of each row
        else:
            patterns = gray_image.copy()
        # print(patterns)
        # print(patterns.shape)
        scaled_data = np.divide(patterns,255)
        
        scaled_data = np.reshape(scaled_data, (-1, 1))
        # print(scaled_data)
        scaled_data = np.insert(scaled_data, 1, 0,axis=1)
        # print(scaled_data)
        # print(">>>>>>>>>>>>>>>>>>>>>>")
        
        self.data = list(self.data)
        self.data.append(scaled_data)
        # print(self.data)
        self.data = np.array(self.data, dtype = object)
        # print(">>>.....>>>>>>>")
        self.kmeans = KMeans(n_clusters = n_clusters, random_state = random_state, max_iter = max_iter)
        self.kmeans.fit(scaled_data)
        
        self.modelLabels = list(self.modelLabels)
        self.modelLabels.append(self.kmeans.labels_)
        # print(self.modelLabels)
        self.modelLabels = np.array(self.modelLabels, dtype = object)
        
        # print(">>>.....>>>>>>>")
        self.relativeDistance.append([])

        self.centroids = list(self.centroids)
        self.centroids.append(self.kmeans.cluster_centers_)
        # print(self.centroids)
        self.centroids = np.array(self.centroids, dtype = object)
        # print(self.centroids)
        # print(self.modelLabels)
        # class1 = np.where(self.modelLabels == 0)[0]
        if self.modelID.shape[0] == 0:
            self.modelID = np.append(self.modelID, 0)
        else:
            self.modelID = np.append(self.modelID, copy.deepcopy(self.modelID[-1]) + 1)
            
    def KMEANSPredictor(self, gray_image:np.ndarray, modelID:int, n_clusters = 2, max_iter = 500, n_init = 1, \
        mode = True, newModel = True) -> np.ndarray:
        
        if mode:
            patterns = np.apply_along_axis(self.__modeGetter, 1, gray_image) # modes of each row
        else:
            patterns = gray_image.copy()
        # print(patterns)
        # print(patterns.shape)
        scaled_data = np.divide(patterns,255)
        scaled_data = np.reshape(scaled_data, (-1, 1))
        # print(scaled_data)
        scaled_data = np.insert(scaled_data, 1, 0,axis=1)
        
        if newModel:
            self.kmeans = KMeans(n_clusters = n_clusters, max_iter = max_iter, init = self.centroids[modelID], \
                n_init = n_init)
            self.kmeans.fit(self.data[modelID])
        labels = self.kmeans.predict(scaled_data)
        
        return labels


class edit_excel():
    """
    1. read_npy(): 读取npy文件，npy文件内为每个page的title，请不要直接打开该文件进行修改。
    2. save_title(): 存入npy文件
    3. new_data_excel(page_title=[], list_Title=[], list_Data=[], member_list=[], mod=0): 进行模式选择
    默认模式为mod=0，即进行”x月_部落战.xlsx“的存入数据,只填写 page_title, list_Title, list_Data 这三个参数;
    当mod=1, 即进行"人员信息统计.xlsx"的存入数据的操作,只填写 member_list和 mod=1 。
    4. open_and_close_txt(): 该函数为try, except, 将报错信息填入txt中。
    5. create_new_folder(),create_new_excel(),create_and_import_sheet()已经集成到了new_data_excel()中，无需
    调用 (mod=0)。
    6. edit_sheet对部落战excel的操作, 未完成
    7. import_member_information(), 已经集成到new_data_excel()中,(mod=1)
    8. edit_member_information(information_edit), 对人员信息进行修改，只填入一个list，注意需要一一对应，并且其中标签
    元素不能被修改。
    9. search_member_information(label)，通过标签来搜索人员全部信息。



    """
    def __init__(self) -> None:
        self.currentPath = os.getcwd()
        self.title_list = []
        self.excel_name = ""
        self.path_now = ""
        self.backUpData = dict()
        self.currentForm = None

    def is_real_number(self,string):
        try:
            float(string)
            return True
        except ValueError:
            return False


    def dataTransferWrapper(self, fileName:str, mapList:list, primaryKeyIndex = [], pageTitle = [],\
        list_Title:list = [], specialDataIndex = []) -> None:
        # dataTitle = list(self.backUpData.keys())
        specialDataIndex = set(specialDataIndex)
        disappearIndex = set()
        
        for x in range(len(self.backUpData[0])):
            dataResult = []
            updateIndexList = []
            keysContent = [self.backUpData[z][x] for z in primaryKeyIndex]
            print("key is", keysContent)
            formData = self.infoSearch(fileName, keysContent)
            if formData != False:
                formData = formData[0]
            print("found data is", formData)
            
            appearIndex = []
            operationList = []
            conditionList = []
            disappearFlag = False
            
            for y in range(len(mapList)):
                print("columns", y)
                if y not in specialDataIndex:
                    dataResult += [self.backUpData[mapList[y]][x]]
                    updateIndexList.append(y)
                else:
                    if mapList[y][0] == "--==//binary operation--==//":
                        firstNum = list_Title.index(mapList[y][1])
                        secondNum = list_Title.index(mapList[y][3])
                        
                        operation = mapList[y][2]
                        print(firstNum, secondNum, operation)
                        lantency = False
                        
                        # if formData == False:
                        if firstNum < len(dataResult) and secondNum < len(dataResult):
                            firstCol = dataResult[firstNum].replace("-", "", 1)
                            secondCol = dataResult[secondNum].replace("-", "", 1)
                            pattern = r"(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2}):(\d{2})"

                            matches1 = re.findall(pattern, dataResult[firstNum])
                            matches2 = re.findall(pattern, dataResult[secondNum])
                            
                            colData1 = dataResult[firstNum]
                            colData2 = dataResult[secondNum]
                        
                        elif firstNum not in specialDataIndex and secondNum not in specialDataIndex:
                            firstCol = self.backUpData[[mapList[firstNum]][x]].replace("-", "", 1)
                            secondCol = self.backUpData[[mapList[secondNum]][x]].replace("-", "", 1)
                            pattern = r"(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2}):(\d{2})"

                            matches1 = re.findall(pattern, self.backUpData[[mapList[firstNum]][x]])
                            matches2 = re.findall(pattern, self.backUpData[[mapList[secondNum]][x]])
                            
                            colData1 = self.backUpData[[mapList[firstNum]][x]]
                            colData2 = self.backUpData[[mapList[secondNum]][x]]
                    
                        else:
                            operationList.append(y)
                            lantency = True
                        # else:
                        #     firstCol = formData[firstNum].replace("-", "", 1)
                        #     secondCol = formData[secondNum].replace("-", "", 1)
                        #     pattern = r"(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2}):(\d{2})"
    
                        #     matches1 = re.findall(pattern, formData[firstNum])
                        #     matches2 = re.findall(pattern, formData[secondNum])
                            
                        #     colData1 = formData[firstNum]
                        #     colData2 = formData[secondNum]
                            
                        print(lantency)
                        print(colData1, colData2)
                        if not lantency:
                            if self.is_real_number(colData1) and self.is_real_number(colData2):
                                if operation == "-":
                                    dataResult += [float(colData1) - float(colData2)]
                                    updateIndexList.append(y)
                                elif operation == "+":
                                    dataResult += [float(colData1) + float(colData2)]
                                    updateIndexList.append(y)
                                elif operation == "x":
                                    dataResult += [float(colData1) * float(colData2)]
                                    updateIndexList.append(y)
                                elif operation == "÷":
                                    dataResult += [float(colData1) / float(colData2)]
                                    updateIndexList.append(y)
                                else:
                                    dataResult += ["placeHolder"]

                            elif len(matches1) == 1 and len(matches2) == 1:
                                if operation == "-":
                                    time1 = datetime.strptime(colData1, \
                                        "%Y-%m-%d %H:%M:%S")
                                    time2 = datetime.strptime(colData2, \
                                        "%Y-%m-%d %H:%M:%S")

                                    # Calculate the time difference
                                    time_diff = time1 - time2

                                    # Extract the desired units
                                    timeResult = str(time_diff)
                                    print(timeResult)
                                    if "day" not in timeResult:
                                        timeResult = "0 days, " + timeResult

                                    dataResult += [timeResult]
                                    updateIndexList.append(y)
                                else:
                                    dataResult += ["placeHolder"]
                            else:
                                dataResult += ["placeHolder"]
                        else:
                            if formData != False:
                                dataResult += [formData[y]]
                            else:
                                dataResult += ["placeHolder"]
                        
                    elif mapList[y][0] == "--==//time--==//":
                        if mapList[y][1] == "Time appear on list YY:MM:DD:H:M:S":
                            appearIndex.append(y)
                            if formData == False:
                                appear = False
                            else:
                                if formData[y] == "None":
                                    appear = False
                                else:
                                    appear = True
                            
                            if not appear:
                                currentTime = time.localtime(time.time())
                                year = currentTime.tm_year  # last two digits of the year
                                month = currentTime.tm_mon
                                day = currentTime.tm_mday
                                hours = currentTime.tm_hour
                                minutes = currentTime.tm_min
                                seconds = currentTime.tm_sec
                            
                                # Format the time components as YY-MM-DD-Hours-Minutes-Seconds (int)
                                formatted_time = "{:02d}-{:02d}-{:02d} {:02d}:{:02d}:{:02d}".format(year, month, day, hours, minutes, seconds)
                                dataResult += [formatted_time]
                                updateIndexList.append(y)
                            else:
                                if formData != False:
                                    dataResult += [formData[y]]
                                else:
                                    dataResult += ["placeHolder"]

                        elif mapList[y][1] == "Time disappear from list YY:MM:DD:H:M:S":
                            if formData != False:
                                dataResult += [formData[y]]
                            else:
                                dataResult += ["placeHolder"]
                                
                            if y not in disappearIndex:
                                disappearIndex.add(y)
                                disappearFlag = True
                            
                    elif mapList[y][0] == "--==//special condition--==//":
                        if mapList[y][1] not in list_Title:
                            firstNum = None
                        else:
                            firstNum = list_Title.index(mapList[y][1])
                            
                        if mapList[y][3] not in list_Title:
                            secondNum = None
                        else:
                            secondNum = list_Title.index(mapList[y][3])
                        
                        operation = mapList[y][2]
                        lantency = False
                        
                        if firstNum == None:
                            firstCol = mapList[y][1].replace("-", "", 1)
                            pattern = r"(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2}):(\d{2})"

                            matches1 = re.findall(pattern, mapList[y][1])
                            
                            if len(matches1) == 0:
                                pattern = r"(-?\d+) *days*, *(\d+):(\d+):(\d+)"

                                matches1 = re.findall(pattern, mapList[y][1])
                                if len(matches1) > 0:
                                    colData1 = int(matches1[0][0])*24*60*60 + int(matches1[0][1])*60*60 + int(matches1[0][2])*60 + int(matches1[0][3])
                                else:
                                    colData1 = mapList[y][1]
                            else:
                                time1 = datetime.strptime(mapList[y][1], "%Y-%m-%d %H:%M:%S")
                                colData1 = time1.timestamp()
                            
                            
                        elif firstNum < len(dataResult):
                            firstCol = dataResult[firstNum].replace("-", "", 1)

                            pattern = r"(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2}):(\d{2})"

                            matches1 = re.findall(pattern, dataResult[firstNum])
                            if len(matches1) == 0:
                                pattern = r"(-?\d+) *days*, *(\d+):(\d+):(\d+)"

                                matches1 = re.findall(pattern, dataResult[firstNum])
                                print(matches1)
                                if len(matches1) > 0:
                                    colData1 = int(matches1[0][0])*24*60*60 + int(matches1[0][1])*60*60 + int(matches1[0][2])*60 + int(matches1[0][3])
                                else:
                                    colData1 = dataResult[firstNum]
                            else:
                                time1 = datetime.strptime(dataResult[firstNum], "%Y-%m-%d %H:%M:%S")
                                colData1 = time1.timestamp()
                                
                            
                        else:
                            lantency = True
                            conditionList.append(y)
                            
                        if secondNum == None:
                            secondCol = mapList[y][3].replace("-", "", 1)
                            pattern = r"(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2}):(\d{2})"
                            matches2 = re.findall(pattern, mapList[y][3])
                            
                            if len(matches2) == 0:
                                pattern = r"(-?\d+) *days*, *(\d+):(\d+):(\d+)"

                                matches2 = re.findall(pattern, mapList[y][3])
                                if len(matches2) > 0:
                                    colData2 = int(matches2[0][0])*24*60*60 + int(matches2[0][1])*60*60 + int(matches2[0][2])*60 + int(matches2[0][3])
                                else:
                                    colData2 = mapList[y][3]
                            else:
                                time2 = datetime.strptime(mapList[y][3], "%Y-%m-%d %H:%M:%S")
                                colData2 = time2.timestamp()
                            
                        elif secondNum < len(dataResult):
                            secondCol = dataResult[secondNum].replace("-", "", 1)
                            pattern = r"(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2}):(\d{2})"
                            matches2 = re.findall(pattern, dataResult[secondNum])
                            if len(matches2) == 0:
                                pattern = r"(-?\d+) *days*, *(\d+):(\d+):(\d+)"

                                matches2 = re.findall(pattern, dataResult[secondNum])
                                if len(matches2) > 0:
                                    colData2 = int(matches2[0][0])*24*60*60 + int(matches2[0][1])*60*60 + int(matches2[0][2])*60 + int(matches2[0][3])
                                else:
                                    colData2 = dataResult[secondNum]
                            else:
                                time2 = datetime.strptime(dataResult[secondNum], "%Y-%m-%d %H:%M:%S")
                                colData2 = time2.timestamp()
                            
                        else:
                            if not lantency:
                                lantency = True
                                conditionList.append(y)
                            
                        codition = False
                        print(colData1, colData2)
                        print(operation)
                        if not lantency:
                            if self.is_real_number(colData1) and self.is_real_number(colData2):
                                if operation == ">":
                                    codition = float(colData1) > float(colData2)
                                elif operation == ">=":
                                    codition = float(colData1) >= float(colData2)
                                    print(float(colData1) >= float(colData2))
                                elif operation == "<":
                                    codition = float(colData1) < float(colData2)
                                elif operation == "<=":
                                    codition = float(colData1) <= float(colData2)
                            elif operation == "==":
                                codition = colData1 == colData2
                            
                            print(codition)
                            if codition:
                                if mapList[y][4] == "Do nothing":
                                    if formData != False:
                                        dataResult += [formData[y]]
                                    else:
                                        dataResult += ["placeHolder"]
                                elif mapList[y][4] == "--==//time--==//":
                                    currentTime = time.localtime(time.time())
                                    year = currentTime.tm_year  # last two digits of the year
                                    month = currentTime.tm_mon
                                    day = currentTime.tm_mday
                                    hours = currentTime.tm_hour
                                    minutes = currentTime.tm_min
                                    seconds = currentTime.tm_sec
                                
                                    # Format the time components as YY-MM-DD-Hours-Minutes-Seconds (int)
                                    formatted_time = "{:02d}-{:02d}-{:02d}-{:02d}-{:02d}-{:02d}".format(year, month, day, hours, minutes, seconds)
                                    dataResult += [formatted_time]
                                    updateIndexList.append(y)
                                else:
                                    dataResult += [mapList[y][4]]
                                    updateIndexList.append(y)
                            else:
                                if mapList[y][5] == "Do nothing":
                                    if formData != False:
                                        dataResult += [formData[y]]
                                    else:
                                        dataResult += ["placeHolder"]
                                elif mapList[y][5] == "--==//time--==//":
                                    currentTime = time.localtime(time.time())
                                    year = currentTime.tm_year  # last two digits of the year
                                    month = currentTime.tm_mon
                                    day = currentTime.tm_mday
                                    hours = currentTime.tm_hour
                                    minutes = currentTime.tm_min
                                    seconds = currentTime.tm_sec
                                
                                    # Format the time components as YY-MM-DD-Hours-Minutes-Seconds (int)
                                    formatted_time = "{:02d}-{:02d}-{:02d}-{:02d}-{:02d}-{:02d}".format(year, month, day, hours, minutes, seconds)
                                    dataResult += [formatted_time]
                                    updateIndexList.append(y)
                                else:
                                    dataResult += [mapList[y][5]]
                                    updateIndexList.append(y)
                        else:
                            if formData != False:
                                dataResult += [formData[y]]
                            else:
                                dataResult += ["placeHolder"]
                                
                print("current data is", dataResult)            
                print("operation is", operationList)
                print("condirion is", conditionList)
            if disappearFlag:
                self.disappearanceCheck(fileName, primaryKeyIndex, list(disappearIndex), appearIndex)
            if len(operationList) > 0:
                for y in operationList:
                    firstNum = list_Title.index(mapList[y][1])
                    secondNum = list_Title.index(mapList[y][3])

                    operation = mapList[y][2]

                    # if formData == False:
                    if firstNum < len(dataResult) and secondNum < len(dataResult):
                        firstCol = dataResult[firstNum].replace("-", "", 1)
                        secondCol = dataResult[secondNum].replace("-", "", 1)
                        pattern = r"(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2}):(\d{2})"

                        matches1 = re.findall(pattern, dataResult[firstNum])
                        matches2 = re.findall(pattern, dataResult[secondNum])

                        colData1 = dataResult[firstNum]
                        colData2 = dataResult[secondNum]

                    elif firstNum not in specialDataIndex and secondNum not in specialDataIndex:
                        firstCol = self.backUpData[[mapList[firstNum]][x]].replace("-", "", 1)
                        secondCol = self.backUpData[[mapList[secondNum]][x]].replace("-", "", 1)
                        pattern = r"(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2}):(\d{2})"

                        matches1 = re.findall(pattern, self.backUpData[[mapList[firstNum]][x]])
                        matches2 = re.findall(pattern, self.backUpData[[mapList[secondNum]][x]])

                        colData1 = self.backUpData[[mapList[firstNum]][x]]
                        colData2 = self.backUpData[[mapList[secondNum]][x]]
                    if self.is_real_number(firstCol) and self.is_real_number(secondCol):
                        if operation == "-":
                            dataResult[y] = float(colData1) - float(colData2)
                            updateIndexList.append(y)
                        elif operation == "+":
                            dataResult[y] = float(colData1) + float(colData2)
                            updateIndexList.append(y)
                        elif operation == "x":
                            dataResult[y] = float(colData1) * float(colData2)
                            updateIndexList.append(y)
                        elif operation == "÷":
                            dataResult[y] = float(colData1) / float(colData2)
                            updateIndexList.append(y)
                    #
                    elif len(matches1) == 1 and len(matches2) == 1:
                        if operation == "-":
                            time1 = datetime.strptime(colData1, \
                                "%Y-%m-%d %H:%M:%S")
                            time2 = datetime.strptime(colData2, \
                                "%Y-%m-%d %H:%M:%S")
                    #
                            # Calculate the time difference
                            time_diff = time1 - time2
                            # Extract the desired units
                            timeResult = str(time_diff)
                            print(timeResult)
                            if "day" not in timeResult:
                                timeResult = "0 days, " + timeResult
                    #
                            dataResult[y] = timeResult
                            updateIndexList.append(y)
                            
            if len(conditionList) > 0:
                print(conditionList)
                for y in conditionList:
                    if mapList[y][1] not in list_Title:
                        firstNum = None
                    else:
                        firstNum = list_Title.index(mapList[y][1])

                    if mapList[y][3] not in list_Title:
                        secondNum = None
                    else:
                        secondNum = list_Title.index(mapList[y][3])


                    if firstNum == None:
                        firstCol = mapList[y][1].replace("-", "", 1)
                        pattern = r"(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2}):(\d{2})"
                        #
                        matches1 = re.findall(pattern, mapList[y][1])

                        if len(matches1) == 0:
                            pattern = r"(-?\d+) *days*, *(\d+):(\d+):(\d+)"
                                #
                            matches1 = re.findall(pattern, mapList[y][1])
                            if len(matches1) > 0:
                                colData1 = int(matches1[0][0])*24*60*60 + int(matches1[0][1])*60*60 + int(matches1[0][2])*60 + int(matches1[0][3])
                            else:
                                colData1 = mapList[y][1]
                        else:
                            time1 = datetime.strptime(mapList[y][1], "%Y-%m-%d %H:%M:%S")
                            colData1 = time1.timestamp()


                    elif firstNum < len(dataResult):
                        firstCol = dataResult[firstNum].replace("-", "", 1)
                            #
                        pattern = r"(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2}):(\d{2})"
                    #
                        matches1 = re.findall(pattern, dataResult[firstNum])
                        if len(matches1) == 0:
                            pattern = r"(-?\d+) *days*, *(\d+):(\d+):(\d+)"
                    #
                            matches1 = re.findall(pattern, dataResult[firstNum])
                            if len(matches1) > 0:
                                colData1 = int(matches1[0][0])*24*60*60 + int(matches1[0][1])*60*60 + int(matches1[0][2])*60 + int(matches1[0][3])
                            else:
                                colData1 = dataResult[firstNum]
                        else:
                            time1 = datetime.strptime(dataResult[firstNum], "%Y-%m-%d %H:%M:%S")
                            colData1 = time1.timestamp()



                    if secondNum == None:
                        secondCol = mapList[y][3].replace("-", "", 1)
                        pattern = r"(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2}):(\d{2})"
                        matches2 = re.findall(pattern, mapList[y][3])

                        if len(matches2) == 0:
                            pattern = r"(-?\d+) *days*, *(\d+):(\d+):(\d+)"
                        #
                            matches2 = re.findall(pattern, mapList[y][3])
                            if len(matches2) > 0:
                                colData2 = int(matches2[0][0])*24*60*60 + int(matches2[0][1])*60*60 + int(matches2[0][2])*60 + int(matches2[0][3])
                            else:
                                colData2 = mapList[y][3]
                        else:
                            time2 = datetime.strptime(mapList[y][3], "%Y-%m-%d %H:%M:%S")
                            colData2 = time2.timestamp()

                    elif secondNum < len(dataResult):
                        secondCol = dataResult[secondNum].replace("-", "", 1)
                        pattern = r"(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2}):(\d{2})"
                        matches2 = re.findall(pattern, dataResult[secondNum])
                        if len(matches2) == 0:
                            pattern = r"(-?\d+) *days*, *(\d+):(\d+):(\d+)"
                        #
                            matches2 = re.findall(pattern, dataResult[secondNum])
                            if len(matches2) > 0:
                                colData2 = int(matches2[0][0])*24*60*60 + int(matches2[0][1])*60*60 + int(matches2[0][2])*60 + int(matches2[0][3])
                            else:
                                colData2 = dataResult[secondNum]
                        else:
                            time2 = datetime.strptime(dataResult[secondNum], "%Y-%m-%d %H:%M:%S")
                            colData2 = time2.timestamp()


                    codition = False
                    if self.is_real_number(firstCol) and self.is_real_number(secondCol):
                        if operation == ">":
                            codition = float(colData1) > float(colData2)
                        elif operation == ">=":
                            codition = float(colData1) >= float(colData2)
                        elif operation == "<":
                            codition = float(colData1) < float(colData2)
                        elif operation == "<=":
                            codition = float(colData1) <= float(colData2)
                    elif operation == "==":
                        codition = colData1 == colData2

                    if codition:
                        if mapList[y][4] == "Do nothing":
                            if formData != False:
                                dataResult[y] = formData[y]
                            else:
                                dataResult[y] = "placeHolder"
                        elif mapList[y][4] == "--==//time--==//":
                            currentTime = time.localtime(time.time())
                            year = currentTime.tm_year  # last two digits of the year
                            month = currentTime.tm_mon
                            day = currentTime.tm_mday
                            hours = currentTime.tm_hour
                            minutes = currentTime.tm_min
                            seconds = currentTime.tm_sec

                            # Format the time components as YY-MM-DD-Hours-Minutes-Seconds (int)
                            formatted_time = "{:02d}-{:02d}-{:02d}-{:02d}-{:02d}-{:02d}".format(year, month, day, hours, minutes, seconds)
                            dataResult[y] = formatted_time
                            updateIndexList.append(y)
                        else:
                            dataResult[y] = mapList[y][4]
                            updateIndexList.append(y)
                    else:
                        if mapList[y][5] == "Do nothing":
                            if formData != False:
                                dataResult[y] = formData[y]
                            else:
                                print("current data is", dataResult)
                                dataResult[y] = "placeHolder"
                        elif mapList[y][5] == "--==//time--==//":
                            currentTime = time.localtime(time.time())
                            year = currentTime.tm_year  # last two digits of the year
                            month = currentTime.tm_mon
                            day = currentTime.tm_mday
                            hours = currentTime.tm_hour
                            minutes = currentTime.tm_min
                            seconds = currentTime.tm_sec

                            # Format the time components as YY-MM-DD-Hours-Minutes-Seconds (int)
                            formatted_time = "{:02d}-{:02d}-{:02d} {:02d}:{:02d}:{:02d}".format(year, month, day, hours, minutes, seconds)
                            dataResult[y] = formatted_time
                            updateIndexList.append(y)
                        else:
                            dataResult[y] = mapList[y][5]
                            updateIndexList.append(y)
            
            for y in list(disappearIndex):
                if formData != False:
                    dataResult[y] = formData[y]
                else:
                    dataResult[y] = "placeHolder"

            if formData != False:
                self.edit_member_information([dataResult])
            else:
                self.new_data_excel(member_list=dataResult, mod=1)
        # for x in specialDataIndex:
        #     if mapList[x][0] == "--==//binary operation--==//":
        #         if mapList[x][1] == "Time appear on list YY:MM:DD:H:M:S":
        #             file_path = os.path.join(self.currentPath, "data\\"+fileName)
        #             if not os.path.exists(file_path):
                        
        #     elif mapList[x][0] == "--==//time--==//":
        #         pass
        #     elif mapList[x][0] == "--==//special condition--==//":
        #         pass
        
    def infoSearch(self, fileName:str, keysList:list):
        file_path = os.path.join(self.currentPath, "data\\"+fileName)
        if not os.path.exists(file_path):
            return False
        result = False
        
        if fileName == "人员信息统计.xlsx":
            result = self.search_member_information([keysList[0]])
            if result == []:
                return False
        
        return result
        
    def appearanceCheck(self, fileName:str, keysList:list):
        file_path = os.path.join(self.currentPath, "data\\"+fileName)
        if not os.path.exists(file_path):
            return False
        
        if fileName == "人员信息统计.xlsx":
            if self.search_member_information([keysList[0]]) == []:
                return False
            
        return True

    def disappearanceCheck(self, fileName:str, keysIndex:list, addTimeLineTo = [], disappearLabelRecord = []):
        file_path = os.path.join(self.currentPath, "data\\"+fileName)
        if not os.path.exists(file_path):
            return False
        
        if fileName == "人员信息统计.xlsx":
            currentExistKeys = np.array(self.backUpData[keysIndex[0]])
            formLabel = np.array(self.allMemberLabel())
            filtered_array = np.setdiff1d(formLabel, currentExistKeys)
            if len(addTimeLineTo) > 0:
                formData = self.search_member_information(list(filtered_array))
                currentTime = time.localtime(time.time())
                year = currentTime.tm_year  # last two digits of the year
                month = currentTime.tm_mon
                day = currentTime.tm_mday
                hours = currentTime.tm_hour
                minutes = currentTime.tm_min
                seconds = currentTime.tm_sec
            
                # Format the time components as YY-MM-DD-Hours-Minutes-Seconds (int)
                formatted_time = "{:02d}-{:02d}-{:02d} {:02d}:{:02d}:{:02d}".format(year, month, day, hours, minutes, seconds)
                for x in range(len(formData)):
                    formData[x][addTimeLineTo[0]] = copy.deepcopy(formatted_time)
                    formData[x][disappearLabelRecord[0]] = "None"
                
                self.edit_member_information(formData)
            
        return True

    def OCRModelDataSaver(self, centroids:np.ndarray, data:np.ndarray, \
        modelID:np.ndarray, relativeDistance:list) ->None :
        """
            `OtherInfo`: [clustersNumber, max_iteration]
        """
        print(centroids, data, modelID, relativeDistance)#, centroids, data, modelID, relativeDistance
        arr = np.array([list(centroids), list(data), list(modelID), list(relativeDistance)], dtype=object)
        
        np.save("textBoxesInfer.npy", arr, allow_pickle=True, fix_imports=True)
        
        print("saved:", arr)

    def OCRModelDataLoader(self) -> np.ndarray:
        file_path = os.path.join(self.currentPath, "textBoxesInfer.npy")

        if os.path.exists(file_path):
            inCodeFile = np.load("textBoxesInfer.npy", allow_pickle = True)

        else:
            inCodeFile = np.array([[], [], [], []], dtype = object)

        
        print("loaded:", inCodeFile)
        return inCodeFile
    
    def filT(self,tem)->bool:
        return os.path.isfile(self.currentPath+"\\UI\\"+tem)
    
    def UIImageList(self) -> list[str]:
        if not os.path.isdir(self.currentPath + "\\UI"):
            os.makedirs(self.currentPath + "\\UI")
            
        return list(filter(self.filT,os.listdir(self.currentPath+"\\UI")))

    def read_npy(self, path):
        if os.path.exists(path):

            the_title_list_np = np.load(path)
            the_title_list = list(the_title_list_np)
            return the_title_list
        else:
            the_title_list = []
            return the_title_list

    def save_title(self, title, path):
        title_np = np.array(title)
        np.save(path, title_np)

    def new_data_excel(self, page_title=[], list_Title=[], list_Data=[], member_list=[], mod=0):
        try:
            if mod == 0:
                self.create_new_folder()
                self.create_new_excel(0)
                self.create_and_import_sheet(page_title, list_Title, list_Data)
            elif mod == 1:
                self.create_new_folder()
                self.create_new_excel(1)
                self.import_member_information(member_list)
        except Exception as e:
            self.open_and_close_txt(e)

    # 打开日志并且填入错误信息
    def open_and_close_txt(self, e):
        localDate = re.sub(r"[ :]+", "-", str(time.asctime(time.localtime(time.time()))))
        f3 = open("log.txt", "a")
        f3.write(localDate + ": " + str(e) + "\n")
        f3.close()

    # 下面三个function要一起调用才是某月数据
    # 创建文件夹（数据库）
    def create_new_folder(self):
        try:
            if not os.path.isdir(self.currentPath + "\\data"):
                os.makedirs(self.currentPath + "\\data")
                self.path_now = self.currentPath + "\\data"
            else:
                self.path_now = self.currentPath + "\\data"
        except Exception as e:
            self.open_and_close_txt(e)


    # name是新excel的名字，请包含完整信息，比如“xxxx.xlsx”，mod默认为0
    def create_new_excel(self, mod=0):
        try:
            if mod == 0:
                current_timestamp = time.time()
                current_time = time.localtime(current_timestamp)
                # 提取当前月份
                current_month = current_time.tm_mon
                name = str(current_month) + "月_部落战.xlsx"
                file_path = os.path.join(self.path_now, name)
                if os.path.exists(file_path):  # 判断该excel是否存在于这个文件夹中
                    self.excel_name = name
                    print("已经存在")
                else:
                    df = pd.DataFrame()
                    df.to_excel(file_path, index=False)
                    self.excel_name = name
            elif mod == 1:
                name = "人员信息统计.xlsx"
                file_path = os.path.join(self.path_now, name)
                if os.path.exists(file_path):  # 判断该excel是否存在于这个文件夹中
                    self.excel_name = name
                    print("已经存在")
                else:
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    sheet['A1'] = '昵称'
                    sheet['B1'] = '标签'
                    sheet['C1'] = '加入'
                    sheet['D1'] = '最近退出'
                    sheet['E1'] = '差值'
                    sheet['F1'] = '常驻认证T/F'
                    workbook.save(file_path)
                    self.excel_name = name
        except Exception as e:
            self.open_and_close_txt(e)

    # 导入数据
    def create_and_import_sheet(self, page_title, list_Title, list_Data):
        try:
            file_path = os.path.join(self.path_now, self.excel_name)
            npy_path = os.path.join(self.path_now, self.excel_name + ".npy")
            print("excel name is: " + self.excel_name)
            self.title_list = self.read_npy(npy_path)

            start_column = 0
            start_column_letter = ""
            actual_page_title_row = 1  # page 的名字的行数
            actual_list_title_row = 2  # 数据的名字的行数

            wb = openpyxl.load_workbook(file_path)  # 展开合并单元格
            ws = wb.active

            if page_title[0] not in self.title_list:  # title_list是page的标题
                self.title_list.append(page_title[0])
                page_number = self.title_list.index(page_title[0])
                start_column = 7 * page_number + 1  # 开始列名的数字
                for i in range(0, len(list_Title)):
                    ws[openpyxl.utils.get_column_letter(start_column + i) + str(actual_page_title_row)] = page_title[0]
                    ws[openpyxl.utils.get_column_letter(start_column + i) + str(actual_list_title_row)] = list_Title[i]
            else:
                page_number = self.title_list.index((page_title[0]))
                start_column = 7 * page_number + 1
            count = 1
            #  判断一行的第一个cell是否为空，如果不是则count加一
            while ws[
                openpyxl.utils.get_column_letter(start_column) + str(actual_list_title_row + count)].value is not None:
                # print(ws[openpyxl.utils.get_column_letter(start_column) + str(actual_list_title_row + count)].value)
                if ws[openpyxl.utils.get_column_letter(start_column) + str(actual_list_title_row + count)].value == \
                        list_Data[0]:
                    # print(count)
                    ws.insert_rows(actual_list_title_row + count)

                    break
                count = count + 1
            #  从这一行的第一个数值开始填写
            for i in range(0, len(list_Data)):
                # print(list_Data[i])
                ws[openpyxl.utils.get_column_letter(start_column + i) + str(actual_list_title_row + count)] \
                    = list_Data[i]
            title_column = 0
            while ws[openpyxl.utils.get_column_letter(1 + title_column * 7) + str(1)].value is not None:
                start_row = 1
                end_row = 1
                start_column = title_column * 7 + 1
                end_column = 7 + title_column * 7
                ws.merge_cells(start_row=start_row, end_row=end_row, start_column=start_column, end_column=end_column)

                title_column += 1
            wb.save(file_path)
            self.save_title(self.title_list, npy_path)
            wb.close()
        except Exception as e:
            self.open_and_close_txt(e)

    # 编辑数据，
    def edit_sheet(self, excel_name, page_title, list_Title, list_Data):
        try:
            file_path = os.getcwd() + "\\data\\" + excel_name[0]
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            col_num = 0
            for column in ws.iter_cols(min_row=1, max_row=1):
                for cell in column:
                    if cell.value == page_title[0]:
                        col_num = cell.column
                        # print(f"The data is in column {col_num}")

            print(col_num)
            target_name = str(list_Data[1])
            the_row = 2
            for row in ws.iter_rows(min_row=2, values_only=True):
                if str(row[col_num]) == target_name:
                    i = col_num
                    for col, data in enumerate(list_Data, start=0):  # 修改整行数据
                        # print(1)
                        cell = ws.cell(row=the_row, column=i)
                        # print(2)
                        cell.value = data
                        i += 1

                    break  # 找到匹配行后跳出循环
                the_row += 1
            wb.save(file_path)
            wb.close()
        except Exception as e:
            self.open_and_close_txt(e)



    # 下面这个function和create_new_folder一起调用，是对人员信息的添加
    def import_member_information(self, member_list):
        try:
            print(member_list)
            name = os.path.join(self.path_now, self.excel_name)
            workbook = openpyxl.load_workbook(name)
            sheet = workbook.active
            new_row = sheet.max_row + 1
            for col, data in enumerate(member_list, start=1):
                sheet.cell(row=new_row, column=col, value=data)
            workbook.save(name)
            workbook.close()
        except Exception as e:
            self.open_and_close_txt(e)

    # 进行人员信息修改
    def edit_member_information(self, the_edit):
        # print(information_edit)
        begin = time.time()
        try:
            path = os.getcwd() + "\\data\\人员信息统计.xlsx"
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active

            for information_edit in the_edit:
                target_name = str(information_edit[1])
                # print(type(target_name))
                target_column = 2
                count = 2
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if str(row[target_column - 1]) == target_name:
                        i = 1
                        for col, data in enumerate(information_edit, start=0):  # 修改整行数据
                            cell = sheet.cell(row=count, column=i)
                            cell.value = data
                            i += 1
                        break  # 找到匹配行后跳出循环
                    count += 1

            workbook.save(path)
            workbook.close()
        except Exception as e:
            self.open_and_close_txt(e)
        print(time.time()-begin)
        
    def search_member_information(self, label):
        try:
            path = os.getcwd() + "\\data\\人员信息统计.xlsx"
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
            the_result = []
            for i in label:
                target_value = i
                target_column = "B"

                target_row = None
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
                    if row[1].value == target_value:
                        target_row = row[1].row
                        break
                if target_row is not None:
                    row_data = [cell.value for cell in sheet[target_row]]
                    the_result.append(row_data)
                    workbook.close()

                else:
                    print("Target value not found in the specified column.")
                    workbook.close()
            return the_result
        except Exception as e:
            self.open_and_close_txt(e)

    def delete_member_information(self, label):
        try:
            path = os.getcwd() + "\\data\\人员信息统计.xlsx"
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
            data_to_delete = label
            for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    if cell.value == data_to_delete:
                        # 找到要删除的数据所在的行
                        sheet.delete_rows(cell.row)
                        break  # 删除完毕，结束查找
            workbook.save(path)
            workbook.close()
        except Exception as e:
            self.open_and_close_txt(e)
            
    def allMemberLabel(self):
        try:
            path = os.getcwd() + "\\data\\人员信息统计.xlsx"
            df = pd.read_excel(path)
            
            return list(df["标签"])
            
        except Exception as e:
            self.open_and_close_txt(e)
        
        

class UserBehaviourController():
    def __init__(self, mouseCon:mouse_control, mouseDetect:eventMouse, OCRCon:OCRController, \
        dbManagement:edit_excel, mainCon:windowsUI) -> None:
        
        self.recoredBehaviours = []
        self.behavioursInterpreter = {"leftClick":mouseCon.move_and_press_mouse, "clickOnButton":\
            mouseCon.clickOnButton, "timeWait":mainCon.timeWait, "iconDetection": mainCon.waitUntilIconDetected,\
                "loopController":mainCon.loopController, "dataForm":dbManagement.dataTransferWrapper}
        self.oldPointer = len(self.recoredBehaviours)
        
        
    def actionRecord(self, action:str, param, actionID = -1, edit = -1):
        if actionID == -1 or actionID >= len(self.recoredBehaviours):
            self.recoredBehaviours.append([])
            actionID = -1
        
        if edit == -1:
            self.recoredBehaviours[actionID].append((action, param))
            
        else:
            self.recoredBehaviours[actionID][edit] = (action, param)
        
    def actionExcute(self, actionID:int, step = -1):
        if len(self.recoredBehaviours) > 0:
            return self.behavioursInterpreter[self.recoredBehaviours[actionID][step][0]], \
                self.recoredBehaviours[actionID][step][1]
        
        
    def delRecord(self, actionID = -1, newDataDel = True):
        if newDataDel:
            if len(self.recoredBehaviours) > 0 and len(self.recoredBehaviours) > self.oldPointer:
                self.recoredBehaviours.pop(actionID)
        else:
            if len(self.recoredBehaviours) > 0:
                self.recoredBehaviours.pop(actionID)


if __name__ == "__main__":
    # startEvent=eventKeyboard()
    # startEvent.StartListener()
    # # x,y=startEvent.mouseGet()
    # # while x!=0 and y!=0:
    # #     time.sleep(0.1)
    # #     x,y=startEvent.mouseGet()
    # time.sleep(9)
    # startEvent.activeFlagSet(1)
    # time.sleep(9)
    # # startEvent.activeFlagSet(1)
    # # time.sleep(15)
    # startEvent.terminate()
    # wind = windowsUI(True, 0.1, "black")
    # # time.sleep(5)

    # print("a")
    # keyTest=eventKeyboard()
    # # keyTest.shortcutFlag=True
    # keyTest.StartListener()
    # time.sleep(10)
    # print(keyTest.statusGet())
    # keyTest.terminate()
    # con=mouse_control()
    # con.move_and_press_mouse(636,21)
    # con.mouse.press(Button.left)
    # con.smooth(1913,195,500)
    # con.mouse.release(Button.left)
    # ocr=OCRController(os.getcwd())
    # text=ocr.textboxSeekerTrainer({"top": 0, "left": 0, "width": 1920, "height": 1080})
    # print(text)
    test = edit_excel()
    current_time = time.time()

    # Convert the current time to a struct_time object
    time_struct = time.localtime(current_time)

    # Extract the required components from the struct_time object
    year = time_struct.tm_year  # last two digits of the year
    month = time_struct.tm_mon
    day = time_struct.tm_mday
    hours = time_struct.tm_hour
    minutes = time_struct.tm_min
    seconds = time_struct.tm_sec

    # Format the time components as YY-MM-DD-Hours-Minutes-Seconds
    formatted_time = "{:02d}-{:02d}-{:02d} {:02d}:{:02d}:{:02d}".format(year, month, day, 1, minutes, seconds)
    # test.new_data_excel(member_list = ["sk","#YHHJJK",000,1,1,"T"],mod=1)
    test.edit_member_information([["sk111","#YHHJJK","None","2023-09-10 14:57:50","placeHolder","T"],["kakakaTTT","#YHHJJ","None","2023-09-10 14:57:50","placeHolder","T"]])
    # print(test.search_member_information(["#YHHJJK","#YHHJJ"]))
    # print(test.allMemberLabel())
    test.backUpData = {0:["sk111", "kakakaTTT"],1:[5160, 4236] ,2:["#YHHJJ", "#YHHJJK"]}
    test.dataTransferWrapper("人员信息统计.xlsx", [0, 2, ['--==//time--==//', 'Time appear on list YY:MM:DD:H:M:S'], ['--==//time--==//', 'Time disappear from list YY:MM:DD:H:M:S'], ['--==//binary operation--==//', '加入', '-', '最近退出'], ['--==//special condition--==//', '差值', '>=', '2 days, 0:00:00', "F", "Do nothing"]], \
        [2], [], ['昵称', '标签', '加入', '最近退出', '差值', '常驻认证T/F'], [2, 3, 4, 5])

